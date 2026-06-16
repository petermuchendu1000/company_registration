"""
Match Namecheap domains to companies in pipeline Excel and fill the Domain column.

Usage:
    python match_domains.py [--dry-run]

Logic:
  - Normalize company name: lowercase, remove Ltd/Limited/Plc/Cic/etc, strip non-alpha
  - Normalize domain: strip TLD, lowercase, non-alpha only
  - Match if: domain_core is a substring of company_core, or vice versa,
    or Levenshtein-style score > threshold
  - Skips domains that are already matched or have obvious mismatch
"""

from __future__ import annotations

import os
import re
import sys
import argparse
import requests
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl
from check_ip import ensure_ip_whitelisted

BASE_DIR = Path(__file__).parent
EXCEL_PATH = BASE_DIR / "pipeline_output" / "companies_pipeline.xlsx"
API_URL = "https://api.namecheap.com/xml.response"
NS = {"nc": "http://api.namecheap.com/xml.response"}

_STOPWORDS = {"ltd", "limited", "plc", "cic", "llc", "uk", "the", "and", "of",
              "for", "services", "group", "solutions", "holdings", "enterprises",
              "international", "global", "management", "consulting", "associates"}


def _normalize(s: str) -> str:
    s = s.lower()
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    words = [w for w in s.split() if w not in _STOPWORDS and len(w) > 1]
    return " ".join(words)


def _domain_core(domain: str) -> str:
    """Extract the second-level domain label and normalize.

    'swift-plus.co.uk' -> 'swiftplus'
    '51stmargarets.site' -> '51stmargarets'
    """
    # Always use only the first label — everything after the first dot is TLD
    sld = domain.lower().split(".")[0]
    return re.sub(r"[^a-z0-9]", "", sld)


def _score(company_norm: str, dom_core: str) -> float:
    """Return a match score 0-1."""
    if not dom_core:
        return 0.0
    company_core = company_norm.replace(" ", "")
    # Exact substring match
    if dom_core in company_core:
        return 0.9
    # Sequential character match — requires long domain core and high hit rate
    # to avoid false positives on short or common-letter domains
    if len(dom_core) >= 7 and all(c in company_core for c in dom_core):
        pos = 0
        seq = 0
        for ch in dom_core:
            idx = company_core.find(ch, pos)
            if idx >= 0:
                seq += 1
                pos = idx + 1
        ratio = seq / len(dom_core)
        if ratio >= 0.85:
            return ratio
    # Word overlap
    company_words = set(company_norm.split())
    dom_words = set(re.findall(r"[a-z]{3,}", dom_core))
    if dom_words:
        overlap = len(company_words & dom_words) / len(dom_words)
        if overlap >= 0.5:
            return overlap
    return 0.0


def get_domains() -> list[str]:
    params = {
        "ApiUser": os.environ.get("NAMECHEAP_API_USER", "muchendu"),
        "ApiKey": os.environ.get("NAMECHEAP_API_KEY", "f81fa145ea414646babce9f8c98cb860"),
        "UserName": os.environ.get("NAMECHEAP_USERNAME", "muchendu"),
        "ClientIp": os.environ.get("NAMECHEAP_CLIENT_IP", "154.159.252.4"),
        "Command": "namecheap.domains.getList",
        "PageSize": "100",
    }
    r = requests.get(API_URL, params=params, timeout=20)
    root = ET.fromstring(r.text)
    if root.get("Status") == "ERROR":
        errs = [e.text for e in root.findall(".//nc:Error", NS)]
        raise RuntimeError(f"Namecheap API error: {errs}")
    return [d.get("Name") for d in root.findall(".//nc:Domain", NS) if d.get("Name")]


def load_companies(wb) -> list[tuple[int, str, str, str]]:
    """Returns list of (row_idx, company_number, company_name, current_domain)."""
    ws = wb.active
    headers = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]

    def col(name_frag):
        for i, h in enumerate(headers, 1):
            if name_frag in h:
                return i
        return None

    cn_col = col("company number")
    name_col = col("company name")
    domain_col = col("domain") if col("domain") else None
    # Find the exact "Domain" column (not "Domain Status" etc.)
    for i, h in enumerate(headers, 1):
        if h.strip() == "domain":
            domain_col = i
            break

    rows = []
    for r in range(2, ws.max_row + 1):
        cn = str(ws.cell(r, cn_col).value or "").strip()
        name = str(ws.cell(r, name_col).value or "").strip()
        existing = str(ws.cell(r, domain_col).value or "").strip() if domain_col else ""
        if cn and name:
            rows.append((r, cn, name, existing))
    return rows, domain_col


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    # Verify public IP is whitelisted (updates .env automatically if IP changed)
    ensure_ip_whitelisted(verbose=True)

    print("Fetching Namecheap domain list...")
    domains = get_domains()

    print(f"Found {len(domains)} domains: {domains}")

    wb = openpyxl.load_workbook(EXCEL_PATH)
    companies, domain_col = load_companies(wb)
    ws = wb.active

    print(f"\nMatching {len(domains)} domains against {len(companies)} companies...\n")

    matched: dict[str, tuple[str, str, float]] = {}  # domain -> (cn, name, score)

    for domain in domains:
        dom_core = _domain_core(domain)
        best_score = 0.0
        best_row = None
        best_name = ""
        best_cn = ""

        for row_idx, cn, name, existing in companies:
            name_norm = _normalize(name)
            score = _score(name_norm, dom_core)
            if score > best_score:
                best_score = score
                best_row = row_idx
                best_name = name
                best_cn = cn

        if best_score >= 0.5:
            matched[domain] = (best_cn, best_name, best_score)
            print(f"  MATCH  {domain!r:40s} -> {best_cn} {best_name!r}  (score={best_score:.2f})")
        else:
            print(f"  SKIP   {domain!r:40s} (best score={best_score:.2f} for {best_name!r})")

    if args.dry_run:
        print("\n[Dry run — not writing to Excel]")
        return

    # Write matches to Excel
    written = 0
    for domain, (cn, name, score) in matched.items():
        for row_idx, rcn, rname, existing in companies:
            if rcn == cn:
                if not existing:
                    ws.cell(row_idx, domain_col).value = domain
                    written += 1
                    print(f"  Wrote {domain} -> row {row_idx} ({name})")
                else:
                    print(f"  Skipped {domain} -> {name} (already has: {existing})")
                break

    try:
        wb.save(EXCEL_PATH)
        print(f"\nDone. {written} domains written to Excel.")
    except PermissionError:
        print(f"\nERROR: Cannot save — close the Excel file first, then re-run.")
        sys.exit(1)


if __name__ == "__main__":
    main()
