"""
Master pipeline orchestrator.

Stages (run in order, all idempotent):
  1. check  — verify public IP is whitelisted in Namecheap
  2. match  — match registered Namecheap domains to companies in Excel
  3. email  — create dev@domain forwarding for every company domain
  4. build  — generate brand overlays, build APK/AAB, generate all assets

Usage:
    python pipeline.py                          # Run all stages
    python pipeline.py --stage match            # Single stage
    python pipeline.py --stage build            # Build only
    python pipeline.py --company 02591663       # Build one company
    python pipeline.py --skip-build             # Regenerate assets only (no Gradle)
    python pipeline.py --stage build --artifact aab  # AAB only
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
from pathlib import Path

BASE_DIR = Path(__file__).parent

# ── helpers ─────────────────────────────────────────────────────────────────

def _banner(title: str) -> None:
    print(f"\n{'='*60}")
    print(f"  {title}")
    print(f"{'='*60}")


def _load_env() -> None:
    env = BASE_DIR / ".env"
    if env.exists():
        for line in env.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip())


# ── Stage 1: IP check ────────────────────────────────────────────────────────

def stage_check() -> None:
    _banner("STAGE 1 — IP check")
    from check_ip import ensure_ip_whitelisted
    ensure_ip_whitelisted(verbose=True)
    print("  Namecheap API: OK")


# ── Stage 2: Domain matching ─────────────────────────────────────────────────

def stage_match(dry_run: bool = False) -> None:
    _banner("STAGE 2 — Domain matching")
    import requests
    import xml.etree.ElementTree as ET
    import openpyxl
    import re

    EXCEL_PATH = BASE_DIR / "pipeline_output" / "companies_pipeline.xlsx"
    API_URL = "https://api.namecheap.com/xml.response"
    NS = {"nc": "http://api.namecheap.com/xml.response"}

    _STOPWORDS = {
        "ltd", "limited", "plc", "cic", "llc", "uk", "the", "and", "of",
        "for", "services", "group", "solutions", "holdings", "enterprises",
        "international", "global", "management", "consulting", "associates",
    }

    def _normalize(s: str) -> str:
        s = s.lower()
        s = re.sub(r"[^a-z0-9 ]", " ", s)
        words = [w for w in s.split() if w not in _STOPWORDS and len(w) > 1]
        return " ".join(words)

    def _domain_core(domain: str) -> str:
        sld = domain.lower().split(".")[0]
        return re.sub(r"[^a-z0-9]", "", sld)

    def _score(company_norm: str, dom_core: str) -> float:
        if not dom_core:
            return 0.0
        company_core = company_norm.replace(" ", "")
        if dom_core in company_core:
            return 0.9
        if len(dom_core) >= 7 and all(c in company_core for c in dom_core):
            pos = 0; seq = 0
            for ch in dom_core:
                idx = company_core.find(ch, pos)
                if idx >= 0:
                    seq += 1; pos = idx + 1
            if seq / len(dom_core) >= 0.85:
                return seq / len(dom_core)
        company_words = set(company_norm.split())
        dom_words = set(re.findall(r"[a-z]{3,}", dom_core))
        if dom_words:
            overlap = len(company_words & dom_words) / len(dom_words)
            if overlap >= 0.5:
                return overlap
        return 0.0

    base = {
        "ApiUser":  os.environ["NAMECHEAP_API_USER"],
        "ApiKey":   os.environ["NAMECHEAP_API_KEY"],
        "UserName": os.environ["NAMECHEAP_USERNAME"],
        "ClientIp": os.environ["NAMECHEAP_CLIENT_IP"],
    }

    r = requests.get(API_URL, params={**base, "Command": "namecheap.domains.getList", "PageSize": "100"}, timeout=30)
    root = ET.fromstring(r.text)
    domains = [d.get("Name") for d in root.findall(".//nc:Domain", NS) if d.get("Name")]
    print(f"  Namecheap domains: {domains}")

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]

    def _col(name: str):
        for i, h in enumerate(headers, 1):
            if h.strip() == name.lower():
                return i
        return None

    cn_col     = _col("company number")
    name_col   = _col("company name")
    domain_col = _col("domain")

    companies = []
    for row in range(2, ws.max_row + 1):
        cn   = str(ws.cell(row, cn_col).value or "").strip()
        name = str(ws.cell(row, name_col).value or "").strip()
        dom  = str(ws.cell(row, domain_col).value or "").strip() if domain_col else ""
        if cn and name:
            companies.append((row, cn, name, dom))

    written = 0
    for domain in domains:
        dom_core   = _domain_core(domain)
        best_score = 0.0
        best_row   = None
        best_cn    = ""
        best_name  = ""
        for row_idx, cn, name, existing in companies:
            score = _score(_normalize(name), dom_core)
            if score > best_score:
                best_score = score; best_row = row_idx; best_cn = cn; best_name = name

        if best_score >= 0.5:
            print(f"  MATCH  {domain!r:42s} -> {best_cn} {best_name!r}  (score={best_score:.2f})")
            if not dry_run:
                for row_idx, cn, name, existing in companies:
                    if cn == best_cn and not existing:
                        ws.cell(row_idx, domain_col).value = domain
                        # Update in-memory list too so next iteration sees it
                        companies = [
                            (r, c, n, domain if c == best_cn else d)
                            for r, c, n, d in companies
                        ]
                        written += 1
                        break
        else:
            print(f"  SKIP   {domain!r:42s} (best score={best_score:.2f})")

    if dry_run:
        print("\n  [Dry run — not writing to Excel]")
    else:
        try:
            wb.save(EXCEL_PATH)
            print(f"\n  {written} new domain(s) written to Excel.")
        except PermissionError:
            print("\n  ERROR: Close the Excel file and re-run.")
            sys.exit(1)


# ── Stage 3: Dev email setup ─────────────────────────────────────────────────

def stage_email(dry_run: bool = False) -> None:
    _banner("STAGE 3 — dev@ email forwarding")
    import requests
    import xml.etree.ElementTree as ET
    import openpyxl

    EXCEL_PATH = BASE_DIR / "pipeline_output" / "companies_pipeline.xlsx"
    APPS_DIR   = BASE_DIR / "pipeline_output" / "apps"
    API_URL    = "https://api.namecheap.com/xml.response"
    NS         = {"nc": "http://api.namecheap.com/xml.response"}

    def _api():
        return {
            "ApiUser":  os.environ["NAMECHEAP_API_USER"],
            "ApiKey":   os.environ["NAMECHEAP_API_KEY"],
            "UserName": os.environ["NAMECHEAP_USERNAME"],
            "ClientIp": os.environ["NAMECHEAP_CLIENT_IP"],
        }

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    ws = wb.active
    headers = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]

    def _col(name: str):
        for i, h in enumerate(headers, 1):
            if h.strip() == name.lower():
                return i
        return None

    cn_col     = _col("company number")
    domain_col = _col("domain")
    email_col  = _col("assigned email")
    wb.close()

    # Re-open for iteration
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    ws = wb.active
    companies = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[cn_col - 1]:
            continue
        cn     = str(row[cn_col - 1]).strip()
        domain = str(row[domain_col - 1] or "").strip() if domain_col else ""
        email  = str(row[email_col - 1] or "").strip() if email_col else ""
        if cn and domain and email:
            companies.append((cn, domain, email))
    wb.close()

    print(f"  Companies with domains: {len(companies)}")

    for cn, domain, assigned_email in companies:
        dev_email = f"dev@{domain}"
        print(f"\n  [{cn}] {domain}")

        # Get existing forwards
        params = {**_api(), "Command": "namecheap.domains.dns.getEmailForwarding", "DomainName": domain}
        r = requests.get(API_URL, params=params, timeout=30)
        root = ET.fromstring(r.text)

        if root.get("Status") == "ERROR":
            errs = [e.text for e in root.findall(".//nc:Error", NS)]
            if any("DNS" in (e or "") for e in errs):
                print(f"    SKIP — external DNS")
                continue
            print(f"    ERROR — {errs}")
            continue

        existing = [(fw.get("mailbox", ""), fw.text or "") for fw in root.findall(".//nc:Forward", NS)]
        existing = [(mb, ft) for mb, ft in existing if mb and ft]

        if any(mb == "dev" for mb, _ in existing):
            print(f"    OK — {dev_email} already set up")
        else:
            new_fwd = existing + [("dev", assigned_email)]
            if dry_run:
                print(f"    DRY — would create {dev_email} -> {assigned_email}")
            else:
                set_params = {**_api(), "Command": "namecheap.domains.dns.setEmailForwarding", "DomainName": domain}
                for i, (mb, ft) in enumerate(new_fwd, 1):
                    set_params[f"mailbox{i}"]   = mb
                    set_params[f"ForwardTo{i}"] = ft
                sr = requests.get(API_URL, params=set_params, timeout=30)
                sroot = ET.fromstring(sr.text)
                if sroot.get("Status") == "ERROR":
                    errs = [e.text for e in sroot.findall(".//nc:Error", NS)]
                    print(f"    ERROR setting forwarding — {errs}")
                    continue
                print(f"    SET — {dev_email} -> {assigned_email}")

        # Update manifest
        manifest_path = APPS_DIR / cn / "manifest.json"
        if manifest_path.exists():
            data = json.loads(manifest_path.read_text(encoding="utf-8"))
            if data.get("support_email") != dev_email:
                if dry_run:
                    print(f"    DRY — manifest support_email: {data.get('support_email')!r} -> {dev_email!r}")
                else:
                    data["support_email"] = dev_email
                    manifest_path.write_text(json.dumps(data, indent=2), encoding="utf-8")
                    print(f"    UPD — manifest support_email -> {dev_email}")
            else:
                print(f"    OK  — manifest already has {dev_email}")


# ── Stage 4: Director ID (XBINDER) ──────────────────────────────────────────

def stage_director_id(
    company_numbers: list[str] | None = None,
    dry_run: bool = False,
    headless: bool = True,
) -> None:
    _banner("STAGE 4 — Director ID (XBINDER)")

    xbinder_email    = os.environ.get("XBINDER_EMAIL", "").strip()
    xbinder_password = os.environ.get("XBINDER_PASSWORD", "").strip()
    if not xbinder_email or not xbinder_password:
        print("  SKIP — set XBINDER_EMAIL and XBINDER_PASSWORD in .env")
        return

    import openpyxl
    import requests as req_lib
    sys.path.insert(0, str(BASE_DIR))
    from xbinder_auto import (
        login, create_bind, search_bind, download_id_via_ui,
        split_combined_id_image, load_director_profile,
    )

    EXCEL_PATH    = BASE_DIR / "pipeline_output" / "companies_pipeline.xlsx"
    COMPANIES_DIR = BASE_DIR / "pipeline_output" / "companies"
    CH_BASE       = "https://api.company-information.service.gov.uk"
    ch_key        = os.environ.get("COMPANIES_HOUSE_API_KEY", "")

    import re as _re
    def _co_dir(cn: str, name: str) -> Path:
        safe   = _re.sub(r'[\\/:*?"<>|]', "", name).strip()
        folder = f"{cn} - {safe}" if safe else cn
        return COMPANIES_DIR / folder

    # ── Load company numbers from Excel ──────────────────────────────────────
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    ws = wb.active
    headers = [str(ws.cell(1, c).value or "").strip().lower()
               for c in range(1, ws.max_column + 1)]

    def _col(name):
        for i, h in enumerate(headers, 1):
            if h == name.lower():
                return i
        return None

    cn_col   = _col("company number")
    name_col = _col("company name")
    dir_col  = _col("directors")
    nat_col  = _col("director nationalities")
    dob_col  = _col("director dob")          # may be None in older Excel files
    gen_col  = _col("director gender")       # may be None in older Excel files

    if not cn_col:
        print("  ERROR: 'Company Number' column not found in Excel.")
        wb.close()
        return

    candidates: list[tuple[str, str, str, str, str]] = []  # (cn, director_name, dob_str, director_gender, co_name)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[cn_col - 1]:
            continue
        cn = str(row[cn_col - 1]).strip()
        if cn.isdigit() and len(cn) < 8:
            cn = cn.zfill(8)
        if company_numbers and cn not in company_numbers:
            continue
        co_name  = str(row[name_col - 1] or "").strip() if name_col else ""
        id_dir   = _co_dir(cn, co_name) / "director_id"
        if (id_dir / "ID Front.jpeg").exists():
            print(f"  [{cn}] OK — ID already downloaded")
            continue

        dirs = str(row[dir_col - 1] or "").strip() if dir_col else ""
        nats = str(row[nat_col - 1] or "").strip() if nat_col else ""
        dobs = str(row[dob_col - 1] or "").strip() if dob_col else ""
        gens = str(row[gen_col - 1] or "").strip() if gen_col else ""

        if not dirs:
            continue

        # Pick the first Kenyan director
        dir_list  = [d.strip() for d in dirs.split(";") if d.strip()]
        nat_list  = [n.strip() for n in nats.split(";")]
        dob_list  = [d.strip() for d in dobs.split(";")]
        gen_list  = [g.strip() for g in gens.split(";")]

        director_name   = ""
        director_dob    = ""
        director_gender = ""
        for idx, nat in enumerate(nat_list):
            if "kenyan" in nat.lower():
                director_name   = dir_list[idx] if idx < len(dir_list) else ""
                director_dob    = dob_list[idx] if idx < len(dob_list) else ""
                director_gender = gen_list[idx] if idx < len(gen_list) else ""
                break

        if not director_name:
            continue  # no Kenyan director

        candidates.append((cn, director_name, director_dob, director_gender, co_name))
    wb.close()

    # ── Backfill DOB from CH API for any candidate missing it ────────────────
    if not ch_key and any(not dob for _, _, dob, _, _ in candidates):
        print("  WARN — COMPANIES_HOUSE_API_KEY not set; DOB will be randomised for missing entries")

    session_ch = req_lib.Session()
    filled: list[tuple[str, str, str, str, str]] = []
    for cn, director_name, director_dob, director_gender, co_name in candidates:
        if director_dob:
            filled.append((cn, director_name, director_dob, director_gender, co_name))
            continue
        # DOB missing — query CH API, match by surname
        dob_str = ""
        if ch_key:
            try:
                resp = session_ch.get(
                    f"{CH_BASE}/company/{cn}/officers",
                    auth=(ch_key, ""),
                    timeout=20,
                )
                if resp.status_code == 200:
                    target_full = director_name.strip().upper()
                    target_surn = target_full.split(",")[0].strip()
                    surn_fallback = ""
                    for o in resp.json().get("items", []):
                        if o.get("resigned_on"):
                            continue
                        if "director" not in (o.get("officer_role", "") or "").lower():
                            continue
                        dob = o.get("date_of_birth") or {}
                        y, m, day = dob.get("year"), dob.get("month"), dob.get("day")
                        if y and m and day:
                            d_val = f"{int(y):04d}-{int(m):02d}-{int(day):02d}"
                        elif y and m:
                            d_val = f"{int(y):04d}-{int(m):02d}"
                        elif y:
                            d_val = f"{int(y):04d}"
                        else:
                            d_val = ""
                        ch_full = (o.get("name") or "").strip().upper()
                        ch_surn = ch_full.split(",")[0].strip()
                        if ch_full == target_full:
                            dob_str = d_val
                            break
                        if ch_surn == target_surn and not surn_fallback:
                            surn_fallback = d_val
                    if not dob_str:
                        dob_str = surn_fallback
            except Exception as e:
                print(f"  [{cn}] CH API DOB lookup error: {e}")
        filled.append((cn, director_name, dob_str, director_gender, co_name))
        if dob_str:
            print(f"  [{cn}] DOB from CH API: {dob_str}")
        else:
            print(f"  [{cn}] DOB not found — will be randomised by load_director_profile")

    candidates = filled

    if not candidates:
        print("  No companies with Kenyan directors found (or all IDs already downloaded).")
        return

    print(f"\n  {len(candidates)} companies need director IDs")

    if dry_run:
        for cn, name, _, gender, _ in candidates:
            print(f"  [{cn}] DRY — would generate ID for: {name}  gender={gender or '?'}")
        return

    # ── Login to XBINDER once ─────────────────────────────────────────────────
    login_session = req_lib.Session()
    try:
        auth    = login(login_session, xbinder_email, xbinder_password)
        token   = auth["token"]
        user_id = auth.get("user_id")
        print("  XBINDER login: OK")
    except Exception as e:
        print(f"  XBINDER login FAILED: {e}")
        return

    # ── Process each company ──────────────────────────────────────────────────
    for cn, director_name, dob_str, director_gender, co_name in candidates:
        id_dir = _co_dir(cn, co_name) / "director_id"
        if (id_dir / "ID Front.jpeg").exists():
            print(f"  [{cn}] OK — ID already downloaded")
            continue

        # Normalise gender: M/F only; fall back to infer from name if still unknown
        gender_val = director_gender.strip().upper()
        if gender_val not in {"M", "F"}:
            from run_pipeline import _infer_gender as _ig
            gender_val = _ig(director_name)
        if gender_val not in {"M", "F"}:
            gender_val = "M"   # last-resort default

        print(f"\n  [{cn}] Generating ID for: {director_name}  gender={gender_val}")

        # Build profile via temp JSON + load_director_profile (proven pattern)
        profile_data = {
            cn: {
                "company_number": cn,
                "xbinder_profile": {
                    "fname": "",
                    "lname": "",
                    "sname": "",
                    "idnumber": "",
                    "actualdob": dob_str,
                    "gender": gender_val,
                },
                "company_data": {
                    "primary_director": {
                        "name": director_name,
                        "date_of_birth": dob_str,
                    }
                },
            }
        }
        temp_profile = BASE_DIR / "pipeline_output" / f"_tmp_profile_{cn}.json"
        try:
            temp_profile.write_text(json.dumps(profile_data), encoding="utf-8")
            profile = load_director_profile(temp_profile, cn)
        except Exception as e:
            print(f"  [{cn}] profile build FAILED: {e}")
            temp_profile.unlink(missing_ok=True)
            continue
        finally:
            temp_profile.unlink(missing_ok=True)

        # Create bind (new session per bind, Bearer-token authenticated)
        try:
            bind_session = req_lib.Session()
            bind_result = create_bind(bind_session, token, profile)
            bind    = bind_result.get("bind", {})
            bind_id = bind.get("id", "")
            id_no   = bind.get("idNo", profile.get("idnumber", ""))
            print(f"  [{cn}] bind: id={bind_id}  id_no={id_no}")
            print(f"  [{cn}] name: {profile['fname']} {profile['sname']} {profile['lname']}")
        except Exception as e:
            print(f"  [{cn}] create_bind FAILED: {e}")
            continue

        if not bind_id:
            print(f"  [{cn}] SKIP — no bind_id in response")
            continue

        # Search confirms bind is indexed (mirrors xbinder_auto main() flow)
        try:
            search_bind(bind_session, token, id_no[:2] if id_no else "")
        except Exception:
            pass

        # Download via browser UI
        id_dir.mkdir(parents=True, exist_ok=True)
        combined = id_dir / "ID_Combined.jpeg"
        try:
            download_id_via_ui(token, user_id, bind_id, id_no, combined, headless=headless)
        except Exception as e:
            print(f"  [{cn}] download FAILED: {e}")
            continue

        # Split into front/back
        try:
            front_path, back_path = split_combined_id_image(combined)
            print(f"  [{cn}] ID done: {front_path.name} + {back_path.name}")
        except Exception as e:
            print(f"  [{cn}] split FAILED (keeping combined): {e}")
            front_path = combined
            back_path  = combined

        # Persist metadata
        result = {
            "company_number":  cn,
            "director_name":   director_name,
            "full_name_on_id": f"{profile.get('fname','')} {profile.get('sname','')} {profile.get('lname','')}".strip(),
            "id_type":         "National ID",
            "id_country":      "KE",
            "bind_id":         bind_id,
            "id_image_front":  str(front_path),
            "id_image_back":   str(back_path),
            "status":          "downloaded",
        }
        (id_dir / "id_result.json").write_text(
            json.dumps(result, indent=2), encoding="utf-8"
        )
        print(f"  [{cn}] Saved id_result.json")


# ── Stage 5: App build ──────────────────────────────────────────────────────

def stage_build(
    company_numbers: list[str] | None = None,
    skip_build: bool = False,
    artifact: str = "both",
) -> None:
    _banner("STAGE 5 — App build")

    sdk = os.environ.get(
        "ANDROID_HOME",
        r"C:\Users\LENOVO\AppData\Local\Android\Sdk",
    )
    java_home = os.environ.get("JAVA_HOME", r"C:\Program Files\Android\Android Studio\jbr")
    os.environ["JAVA_HOME"] = java_home

    sys.path.insert(0, str(BASE_DIR))
    from apps.generator.catalog import load_companies
    from apps.generator.generate import (
        build_flavors, run_gradle, collect_artifacts,
        backup_and_make_assets, _ensure_local_properties,
    )

    companies = load_companies(
        company_numbers=company_numbers,
        limit=None,
    )
    # Only build companies that have a domain
    companies = [c for c in companies if c.domain]
    if not companies:
        print("  No companies with domains — nothing to build.")
        return

    print(f"  Building {len(companies)} company app(s):")
    for c in companies:
        print(f"    {c.company_number}  {c.display_name}  ({c.domain})")

    _ensure_local_properties(sdk)

    built = build_flavors(companies)
    print(f"  Wrote {len(built)} flavor overlays + build.gradle.kts")

    if skip_build:
        print("  Gradle: SKIPPED (--skip-build)")
        from apps.generator.generate import collect_artifacts
        got = collect_artifacts(built, variant="release", artifacts=artifact)
        backup_and_make_assets(built, got)
        return

    rc = run_gradle(built, variant="Release", artifact=artifact)
    if rc != 0:
        print(f"  Gradle FAILED (exit {rc})")
        sys.exit(rc)

    got = collect_artifacts(built, variant="release", artifacts=artifact)
    print(f"  Collected {len(got)}/{len(built)} artifact set(s)")

    backup_and_make_assets(built, got)
    print("  Assets generated.")


# ── Summary ──────────────────────────────────────────────────────────────────

def print_summary() -> None:
    _banner("PIPELINE SUMMARY")
    apps_dir = BASE_DIR / "pipeline_output" / "apps"
    if not apps_dir.exists():
        print("  No pipeline_output/apps directory found.")
        return

    for cn_dir in sorted(apps_dir.iterdir()):
        if not cn_dir.is_dir():
            continue
        manifest_path = cn_dir / "manifest.json"
        if not manifest_path.exists():
            continue
        m = json.loads(manifest_path.read_text(encoding="utf-8-sig"))
        apk_ok  = "OK" if m.get("apk") and (cn_dir / m["apk"]).exists() else "--"
        aab_ok  = "OK" if m.get("aab") and (cn_dir / m["aab"]).exists() else "--"
        pp_ok   = "OK" if m.get("privacy_policy_url") else "--"
        dom_ok  = "OK" if m.get("domain") else "--"
        id_ok   = "OK" if (BASE_DIR / "pipeline_output" / "representative_ids" / cn_dir.name / "ID Front.jpeg").exists() else "--"
        email   = m.get("support_email", "")
        print(f"  {cn_dir.name}  APK:{apk_ok}  AAB:{aab_ok}  Policy:{pp_ok}  Domain:{dom_ok}  ID:{id_ok}  email={email}")


# ── Entry point ──────────────────────────────────────────────────────────────

STAGES = ["check", "match", "email", "director_id", "build"]


def main() -> None:
    p = argparse.ArgumentParser(description="Company app pipeline")
    p.add_argument("--stage",       choices=STAGES, help="Run a single stage only")
    p.add_argument("--company",     nargs="+", metavar="CN", help="Limit build to these company numbers")
    p.add_argument("--skip-build",  action="store_true", help="Skip Gradle (regenerate assets only)")
    p.add_argument("--artifact",    default="both", choices=["apk", "aab", "both"])
    p.add_argument("--dry-run",     action="store_true", help="Preview changes without writing")
    p.add_argument("--no-headless", action="store_true", help="Show browser window during XBINDER download (debug)")
    args = p.parse_args()

    _load_env()

    run = args.stage or "all"
    t0  = time.time()

    if run in ("all", "check"):
        stage_check()

    if run in ("all", "match"):
        stage_match(dry_run=args.dry_run)

    if run in ("all", "email"):
        stage_email(dry_run=args.dry_run)

    if run in ("all", "director_id"):
        stage_director_id(
            company_numbers=args.company,
            dry_run=args.dry_run,
            headless=not args.no_headless,
        )

    if run in ("all", "build"):
        stage_build(
            company_numbers=args.company,
            skip_build=args.skip_build,
            artifact=args.artifact,
        )

    print_summary()
    print(f"\n  Total time: {time.time() - t0:.1f}s")


if __name__ == "__main__":
    main()
