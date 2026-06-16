"""
Per-company stage control API — v2.
Register with:  from api_v2 import api_v2; app.register_blueprint(api_v2)
"""
from __future__ import annotations

import io, json, os, re, shutil, sys, threading, time, traceback
from datetime import datetime
from pathlib import Path

import openpyxl
from flask import Blueprint, jsonify, request, send_file, abort

api_v2 = Blueprint("api_v2", __name__)

ROOT     = Path(__file__).parent
EXCEL    = ROOT / "pipeline_output" / "companies_pipeline.xlsx"
COMP_DIR = ROOT / "pipeline_output" / "companies"
DL_BASE  = ROOT / "pipeline_output"          # files must be under here to download

STAGE_ORDER = ["details", "duns", "certificate", "domain", "email", "director_id", "app"]

# ── Job state ──────────────────────────────────────────────────────────────
_JOBS: dict[str, dict] = {}   # key: "{cn8}:{stage}"
_LOCK = threading.Lock()
_CO_LOCKS: dict[str, threading.Lock] = {}   # per-company serialisation


def _co_lock(cn: str) -> threading.Lock:
    with _LOCK:
        if cn not in _CO_LOCKS:
            _CO_LOCKS[cn] = threading.Lock()
        return _CO_LOCKS[cn]


def _jk(cn: str, stage: str) -> str:
    return f"{cn}:{stage}"


def _job_get(cn: str, stage: str) -> dict:
    with _LOCK:
        j = dict(_JOBS.get(_jk(cn, stage), {}))
    if j.get("status") == "running":
        live = _TEE.get_log(_jk(cn, stage))
        if live:
            j["log"] = live
    return j


def _job_set(cn: str, stage: str, **kw) -> None:
    with _LOCK:
        _JOBS.setdefault(_jk(cn, stage), {}).update(kw)


# ── Stdout capture ─────────────────────────────────────────────────────────
# Single global _TEE is installed at sys.stdout once. Each worker thread calls
# _TEE.bind(key) so its prints are routed to its own buffer; no global
# sys.stdout replacement, no race condition between concurrent jobs.

_TEE_LOCK = threading.Lock()


class _Tee:
    """Global stdout multiplexer — installed once at module load."""

    def __init__(self):
        self._real = sys.__stdout__
        self._local = threading.local()
        self._logs: dict[str, list[str]] = {}  # key → chunks

    def bind(self, key: str) -> None:
        self._local.key = key
        with _TEE_LOCK:
            self._logs[key] = []

    def get_log(self, key: str) -> str:
        with _TEE_LOCK:
            return "".join(self._logs.get(key, []))

    def unbind(self) -> str:
        key = getattr(self._local, "key", None)
        log = ""
        if key:
            with _TEE_LOCK:
                log = "".join(self._logs.pop(key, []))
            self._local.key = None
        return log

    def write(self, s: str) -> int:
        if s:
            try: self._real.write(s); self._real.flush()
            except Exception: pass
            key = getattr(self._local, "key", None)
            if key:
                with _TEE_LOCK:
                    if key in self._logs:
                        self._logs[key].append(s)
        return len(s) if s else 0

    def flush(self):
        try: self._real.flush()
        except Exception: pass


_TEE = _Tee()
sys.stdout = _TEE


def _spawn(cn: str, stage: str, fn) -> None:
    """Run fn() in a daemon thread, tracking job state and capturing stdout."""
    key = _jk(cn, stage)
    _job_set(cn, stage, status="running", started_at=datetime.now().isoformat(),
             finished_at=None, error=None, log=None)

    def _worker():
        _TEE.bind(key)
        try:
            result = fn()
            log = _TEE.unbind()
            _job_set(cn, stage, status="done",
                     finished_at=datetime.now().isoformat(),
                     log=log, result=str(result)[:500])
        except Exception as e:
            log = _TEE.unbind()
            _job_set(cn, stage, status="failed",
                     finished_at=datetime.now().isoformat(),
                     error=str(e),
                     log=log + "\n" + traceback.format_exc())

    threading.Thread(target=_worker, daemon=True).start()


# ── Excel helpers ──────────────────────────────────────────────────────────

def _cn8(v) -> str:
    return str(v or "").strip().zfill(8)


def _safe(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "", name).strip()


def _co_dir(cn: str, name: str = "") -> Path:
    folder = f"{_cn8(cn)} - {_safe(name)}" if name else _cn8(cn)
    return COMP_DIR / folder


_ROWS_CACHE: tuple[float, list[dict]] | None = None  # (mtime, rows)


def _all_rows() -> list[dict]:
    global _ROWS_CACHE
    if not EXCEL.exists():
        return []
    try:
        mtime = EXCEL.stat().st_mtime
        if _ROWS_CACHE and _ROWS_CACHE[0] == mtime:
            return _ROWS_CACHE[1]
        wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        hdr_row = next(rows_iter, None)
        if not hdr_row:
            wb.close()
            return []
        hdr = [str(v or "").strip() for v in hdr_row]
        rows = []
        for row in rows_iter:
            if not row or not row[0]: continue
            rows.append({hdr[i]: (row[i] if i < len(row) else None) for i in range(len(hdr))})
        wb.close()
        _ROWS_CACHE = (mtime, rows)
        return rows
    except Exception:
        return []


def _find_row(cn: str) -> dict | None:
    for r in _all_rows():
        if _cn8(r.get("Company Number")) == _cn8(cn):
            return r
    return None


def _update_excel_row(cn: str, updates: dict) -> None:
    if not EXCEL.exists(): return
    try:
        wb = openpyxl.load_workbook(EXCEL)
        ws = wb.active
        hdr = [str(c.value or "").strip() for c in ws[1]]

        def _col(name: str) -> int:
            for i, h in enumerate(hdr):
                if h.lower() == name.lower(): return i + 1
            nc = len(hdr) + 1
            ws.cell(row=1, column=nc, value=name)
            hdr.append(name)
            return nc

        cn_idx = next((i for i, h in enumerate(hdr) if h.lower() == "company number"), None)
        if cn_idx is None: wb.close(); return

        for row in ws.iter_rows(min_row=2):
            if _cn8(row[cn_idx].value) == _cn8(cn):
                for k, v in updates.items():
                    ws.cell(row=row[cn_idx].row, column=_col(k), value=v)
                break

        wb.save(EXCEL)
        wb.close()
    except Exception as e:
        print(f"  [excel-update] {e}")


def _add_company_row(cn: str, name: str) -> None:
    """Insert a new row for a company that isn't in Excel yet."""
    if not EXCEL.exists():
        os.makedirs(EXCEL.parent, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Companies Pipeline"
        from openpyxl.styles import Font, PatternFill, Alignment
        headers = [
            "No.", "Company Number", "Company Name", "Status", "Type",
            "Date of Creation", "SIC Codes", "Address",
            "Directors", "Director Nationalities", "Director DOB", "Director Gender",
            "Buvei First Name", "Buvei Last Name",
            "DUNS Number", "DUNS Status", "DUNS Email Used",
            "Certificate Downloaded", "Certificate Path",
            "Domain", "Domain Status", "Domain Cost",
            "Assigned Email", "Account Name",
        ]
        ws.append(headers)
        fill = PatternFill(start_color="003078", end_color="003078", fill_type="solid")
        font = Font(color="FFFFFF", bold=True, size=11)
        for cell in ws[1]:
            cell.fill = fill; cell.font = font
        wb.save(EXCEL)
        wb.close()

    wb = openpyxl.load_workbook(EXCEL)
    ws = wb.active
    hdr = [str(c.value or "").strip() for c in ws[1]]

    # Check duplicate
    cn_idx = next((i for i, h in enumerate(hdr) if h.lower() == "company number"), None)
    if cn_idx is not None:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and _cn8(row[cn_idx]) == _cn8(cn):
                wb.close(); return  # already exists

    def col(name):
        for i, h in enumerate(hdr):
            if h.lower() == name.lower(): return i + 1
        return None

    nr = ws.max_row + 1
    no = nr - 1
    if col("No."): ws.cell(row=nr, column=col("No."), value=no)
    if col("Company Number"): ws.cell(row=nr, column=col("Company Number"), value=cn)
    if col("Company Name"): ws.cell(row=nr, column=col("Company Name"), value=name)

    wb.save(EXCEL)
    wb.close()


# ── Stage status ───────────────────────────────────────────────────────────

def _compute_stages(cn: str, row: dict) -> dict:
    name  = str(row.get("Company Name") or "")
    cdir  = _co_dir(cn, name)

    has_cert = bool(row.get("Certificate Downloaded"))
    if not has_cert:
        cd = cdir / "certificate"
        has_cert = cd.is_dir() and any(f.endswith("_cert.pdf") for f in os.listdir(cd))

    ad = cdir / "app" / "artifacts"
    has_app = ad.is_dir() and any(f.endswith((".apk", ".aab")) for f in os.listdir(ad))

    return {
        "details":     bool(name),
        "duns":        bool(row.get("DUNS Number")),
        "certificate": has_cert,
        "domain":      bool(row.get("Domain")),
        "email":       bool(row.get("Assigned Email")),
        "director_id": (cdir / "director_id" / "ID Front.jpeg").exists(),
        "app":         has_app,
    }


# ── File helpers ───────────────────────────────────────────────────────────

def _ls(d: Path, exts: list | None = None, names: list | None = None) -> list[dict]:
    if not d.is_dir(): return []
    out = []
    for f in sorted(d.rglob("*")):
        if not f.is_file(): continue
        rel = f.relative_to(d).as_posix()
        if exts and not any(rel.lower().endswith(e) for e in exts): continue
        if names and f.name not in names: continue
        out.append({"name": rel, "size": f.stat().st_size, "path": str(f)})
    return out


def _pp(cdir: Path) -> dict | None:
    for p in [cdir / "payment_profile" / "payment_profile.txt", cdir / "payment_profile.txt"]:
        if p.exists():
            return {"name": "payment_profile.txt", "size": p.stat().st_size, "path": str(p)}
    return None


def _version_file(cn: str, name: str) -> Path:
    return _co_dir(cn, name) / "app" / "version.json"


def _read_version(cn: str, name: str) -> dict:
    vf = _version_file(cn, name)
    if vf.exists():
        try:
            return json.loads(vf.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {"version_code": 1, "version_name": "1.0"}


def _write_version(cn: str, name: str, version_code: int, version_name: str) -> None:
    vf = _version_file(cn, name)
    vf.parent.mkdir(parents=True, exist_ok=True)
    vf.write_text(
        json.dumps({"version_code": version_code, "version_name": version_name}, indent=2),
        encoding="utf-8",
    )


def _read_manifest(cn: str, name: str) -> dict:
    """Read the app's manifest.json; return {} if not found."""
    f = _co_dir(cn, name) / "app" / "manifest.json"
    if f.exists():
        try:
            return json.loads(f.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def _pc_file(cn: str, name: str) -> Path:
    return _co_dir(cn, name) / "app" / "play_console.json"


def _read_pc(cn: str, name: str, application_id: str = "", manifest: dict | None = None,
             assigned_email: str = "", domain: str = "") -> dict:
    """Return saved Play Console data, pre-filled with low-review-risk defaults."""
    if manifest is None:
        manifest = {}

    role_noun        = str(manifest.get("role_noun") or "staff")
    role_verb_start  = str(manifest.get("role_verb_start") or "start")
    role_verb_end    = str(manifest.get("role_verb_end") or "end")
    export_title     = str(manifest.get("export_title") or "Work Log")
    display_name     = str(manifest.get("display_name") or name)
    support_email    = assigned_email or str(manifest.get("support_email") or "")
    domain_val       = domain or str(manifest.get("domain") or "")

    short_desc = f"{display_name[:22]} — {role_noun.capitalize()} shift & hours tracker"[:80]

    full_desc = (
        f"Manage {role_noun} schedules and track working hours with ease.\n\n"
        f"{display_name} provides a simple, reliable solution for UK businesses to "
        f"record {role_noun} shifts, monitor attendance, and export time logs for payroll.\n\n"
        f"Key features:\n"
        f"• Log shift {role_verb_start} and {role_verb_end} times instantly\n"
        f"• View weekly schedule at a glance\n"
        f"• Export {export_title} reports\n"
        f"• Secure and private — data stays on your device\n"
        f"• Designed for UK Working Time Regulations compliance\n\n"
        f"Simple, fast, and built for everyday reliability."
    )

    defaults: dict = {
        # 1. Create App
        "app_name":             name[:30],
        "package_name":         application_id or "",
        "default_language":     "en-GB",
        "app_type":             "app",
        "is_free":              True,
        "policy_confirmed":     False,
        "signing_tos_accepted": False,
        "export_laws_accepted": False,
        # 2. App Content
        "app_access":           "all_available",
        "contains_ads":         False,
        # 3. Content Ratings
        "ratings_category":     "all_other",
        "ratings_email":        support_email,
        "q_violence":           False,
        "q_sexual":             False,
        "q_teen_rating":        False,
        "q_profanity":          False,
        "q_controlled_drugs":   False,
        "q_gambling":           False,
        "q_user_generated":     False,
        "q_account_sharing":    False,
        "q_location_sharing":   False,
        "ratings_iarc_agreed":  False,
        # 4. Target Audience
        "target_age":           "18_and_over",
        "legal_compliance":     False,
        # 5. Data Safety
        "data_collection":      False,
        "families_policy":      False,
        # 6. Policy Declarations
        "is_government":        False,
        "financial_features":   "none",
        "uses_advertising_id":  False,
        "health_features":      "none",
        # 7. Store Listing
        "app_category":         "BUSINESS",
        "contact_email":        support_email,
        "contact_website":      f"https://{domain_val}" if domain_val else "",
        "short_description":    short_desc,
        "full_description":     full_desc[:4000],
        # 8. Release
        "release_name":         "",
        "release_notes":        (
            f"<en-US>\n"
            f"Initial release of {display_name}.\n\n"
            f"• {role_noun.capitalize()} shift {role_verb_start} and {role_verb_end} time logging\n"
            f"• Weekly schedule overview\n"
            f"• Export {export_title} reports\n"
            f"• Designed for UK Working Time Regulations compliance\n"
            f"</en-US>"
        ),
    }

    f = _pc_file(cn, name)
    if f.exists():
        try:
            saved = json.loads(f.read_text(encoding="utf-8"))
            defaults.update(saved)
        except Exception:
            pass

    if application_id and not defaults.get("package_name"):
        defaults["package_name"] = application_id

    return defaults


def _write_pc(cn: str, name: str, data: dict) -> None:
    f = _pc_file(cn, name)
    f.parent.mkdir(parents=True, exist_ok=True)
    f.write_text(json.dumps(data, indent=2), encoding="utf-8")


def _build_pc(cn: str, name: str, app_id: str, manifest: dict,
              assigned_email: str, domain: str) -> dict:
    """Assemble Play Console data, backfilling release_name from version.json."""
    pc = _read_pc(cn, name, app_id, manifest=manifest,
                  assigned_email=assigned_email, domain=domain)
    if not pc.get("release_name"):
        ver = _read_version(cn, name)
        pc["release_name"] = ver.get("version_name", "1.0")
    return pc


def _rmdir(path: Path) -> None:
    """Remove a directory and all its contents, silently."""
    try:
        if path.is_dir():
            shutil.rmtree(path)
    except Exception as e:
        print(f"  [rmdir] {path}: {e}")


def _clean_stage_output(cn: str, name: str, stage: str) -> None:
    """Delete prior output files for a stage before re-running it."""
    cdir = _co_dir(cn, name)
    if stage == "certificate":
        _rmdir(cdir / "certificate")
    elif stage == "director_id":
        _rmdir(cdir / "director_id")
    elif stage == "app":
        arts = cdir / "app" / "artifacts"
        if arts.is_dir():
            for f in arts.iterdir():
                if f.suffix in {".apk", ".aab"}:
                    try: f.unlink()
                    except Exception: pass


def _tree_node(path: Path) -> dict | None:
    """Recursive tree builder (no caching — called by _tree which caches root)."""
    if not path.exists():
        return None
    if path.is_file():
        return {"name": path.name, "type": "file",
                "size": path.stat().st_size, "path": str(path)}
    children: list[dict] = []
    try:
        for child in sorted(path.iterdir(),
                            key=lambda p: (p.is_file(), p.name.lower())):
            node = _tree_node(child)
            if node:
                children.append(node)
    except PermissionError:
        pass
    return {"name": path.name, "type": "dir", "children": children}


_TREE_CACHE: dict[str, tuple[float, dict | None]] = {}  # path → (ts, result)
_TREE_TTL = 10.0  # seconds


def _tree(path: Path) -> dict | None:
    """Return a nested tree dict, cached for _TREE_TTL seconds per root path."""
    key = str(path)
    now = time.monotonic()
    entry = _TREE_CACHE.get(key)
    if entry and now - entry[0] < _TREE_TTL:
        return entry[1]
    result = _tree_node(path)
    _TREE_CACHE[key] = (now, result)
    return result


# ── Retry helper ───────────────────────────────────────────────────────────

def _with_retry(fn, retries: int = 3, delay: float = 2.0):
    """Call fn(); on exception retry up to retries times with exponential backoff."""
    last_exc: Exception | None = None
    for attempt in range(retries):
        try:
            return fn()
        except Exception as e:
            last_exc = e
            if attempt < retries - 1:
                wait = delay * (2 ** attempt)
                print(f"  [retry {attempt + 1}/{retries}] {e} — retrying in {wait:.0f}s")
                time.sleep(wait)
    raise last_exc  # type: ignore[misc]


# ── Core stage executor ────────────────────────────────────────────────────

def _exec_stage(cn: str, stage: str, row: dict):
    """Execute one stage. Called from thread. stdout already redirected."""
    name = str(row.get("Company Name") or "")
    sys.path.insert(0, str(ROOT))

    if stage == "details":
        from run_pipeline import get_company_details
        d = _with_retry(lambda: get_company_details(cn))
        _update_excel_row(cn, {
            "Company Name":           d.get("company_name", ""),
            "Status":                 d.get("company_status", ""),
            "Type":                   d.get("type", ""),
            "Date of Creation":       d.get("date_of_incorporation", ""),
            "SIC Codes":              d.get("sic_codes", ""),
            "Address":                d.get("address", ""),
            "Directors":              d.get("director_names", ""),
            "Director Nationalities": d.get("director_nationalities", ""),
            "Director DOB":           d.get("director_dobs", ""),
            "Director Gender":        d.get("director_genders", ""),
        })
        return d

    elif stage == "duns":
        from run_pipeline import get_duns_number
        r = get_duns_number(cn, headless=True)
        if r.get("duns_number"):
            _update_excel_row(cn, {"DUNS Number": r["duns_number"], "DUNS Status": "found"})
        elif r.get("status") == "submitted":
            _update_excel_row(cn, {"DUNS Status": "submitted",
                                   "DUNS Email Used": r.get("temp_email", "")})
        return r

    elif stage == "certificate":
        _clean_stage_output(cn, name, "certificate")
        from run_pipeline import download_certificate
        out = str(_co_dir(cn, name) / "certificate")
        path, err = _with_retry(lambda: download_certificate(cn, out))
        if path:
            _update_excel_row(cn, {"Certificate Downloaded": True, "Certificate Path": path})
        return {"path": path, "error": err}

    elif stage == "domain":
        from run_pipeline import register_company_domain
        details = {
            "company_name": name, "company_number": cn,
            "sic_codes": str(row.get("SIC Codes") or ""),
            "address": str(row.get("Address") or ""),
        }
        r = register_company_domain(name, details)
        if r.get("domain"):
            _update_excel_row(cn, {"Domain": r["domain"],
                                   "Domain Status": r.get("status", ""),
                                   "Domain Cost": r.get("charged", "")})
        return r

    elif stage == "email":
        from email_pool import EmailPool
        pool = EmailPool()
        ex_email, ex_first, ex_last = pool.get_assigned_email(cn)
        if ex_email:
            # Already assigned in pool — sync to Excel if missing
            if not str(row.get("Assigned Email") or ""):
                _update_excel_row(cn, {"Assigned Email": ex_email,
                                       "Account Name": f"{ex_first} {ex_last}"})
            return {"email": ex_email, "status": "already_assigned"}
        email, first, last = pool.assign_next(company_number=cn, company_name=name)
        _update_excel_row(cn, {"Assigned Email": email,
                                "Account Name": f"{first} {last}"})
        return {"email": email, "status": "assigned"}

    elif stage == "director_id":
        _clean_stage_output(cn, name, "director_id")
        import pipeline as _pl
        _pl.stage_director_id(company_numbers=[cn])
        return {"status": "done"}

    elif stage == "app":
        _clean_stage_output(cn, name, "app")
        import pipeline as _pl
        _pl.stage_build(company_numbers=[cn])
        return {"status": "done"}

    raise ValueError(f"Unknown stage: {stage}")


# ── Routes ─────────────────────────────────────────────────────────────────

@api_v2.route("/app")
def app_page():
    return send_file(ROOT / "static" / "app.html")


@api_v2.route("/api/v2/companies")
def v2_companies():
    rows = _all_rows()
    archived_param = request.args.get("archived", "false").lower()
    show_archived = archived_param == "true"
    
    out = []
    for row in rows:
        cn = _cn8(row.get("Company Number"))
        if not cn or cn == "00000000": continue
        
        # Filter by archive status
        archived_val = str(row.get("Archived") or "").strip().lower()
        is_archived = archived_val in ("yes", "true", "1")
        if is_archived != show_archived:
            continue
        
        stages = _compute_stages(cn, row)
        running = [s for s in STAGE_ORDER if _job_get(cn, s).get("status") == "running"]
        company_status = str(row.get("Company Status") or "").strip() or "Active"
        
        out.append({
            "cn":              cn,
            "name":            str(row.get("Company Name") or ""),
            "domain":          str(row.get("Domain") or ""),
            "stages":          stages,
            "n_done":          sum(stages.values()),
            "running":         running,
            "company_status":  company_status,
            "notes":           str(row.get("Company Notes") or "").strip(),
            "archived":        is_archived,
        })
    return jsonify(out)


@api_v2.route("/api/v2/company/<cn>")
def v2_company(cn):
    row = _find_row(cn)
    if not row:
        return jsonify({"error": "Not found"}), 404

    name     = str(row.get("Company Name") or "")
    cdir     = _co_dir(cn, name)
    stages   = _compute_stages(cn, row)
    manifest    = _read_manifest(cn, name)
    app_id      = str(manifest.get("application_id") or "")
    privacy_url = str(manifest.get("privacy_policy_url") or "")
    a_email     = str(row.get("Assigned Email") or "")
    domain_val  = str(row.get("Domain") or "")

    jobs = {}
    for s in STAGE_ORDER:
        j = _job_get(cn, s)
        if j: jobs[s] = {k: v for k, v in j.items() if k != "log"}

    return jsonify({
        "cn":               cn,
        "name":             name,
        "status":           str(row.get("Status") or ""),
        "type":             str(row.get("Type") or ""),
        "sic":              str(row.get("SIC Codes") or ""),
        "address":          str(row.get("Address") or ""),
        "directors":        str(row.get("Directors") or ""),
        "nationalities":    str(row.get("Director Nationalities") or ""),
        "dobs":             str(row.get("Director DOB") or ""),
        "genders":          str(row.get("Director Gender") or ""),
        "buvei_first":      str(row.get("Buvei First Name") or ""),
        "buvei_last":       str(row.get("Buvei Last Name") or ""),
        "duns":             str(row.get("DUNS Number") or ""),
        "domain":           str(row.get("Domain") or ""),
        "email":            str(row.get("Assigned Email") or ""),
        "account_name":     str(row.get("Account Name") or ""),
        "application_id":   app_id,
        "privacy_policy_url": privacy_url,
        "company_status":   str(row.get("Company Status") or "").strip() or "Active",
        "notes":            str(row.get("Company Notes") or "").strip(),
        "archived":         str(row.get("Archived") or "").strip().lower() in ("yes", "true", "1"),
        "stages":           stages,
        "files": {
            "certificate":   _ls(cdir / "certificate", exts=[".pdf"]),
            "director_id":   _ls(cdir / "director_id",
                                 names=["ID Front.jpeg", "ID Back.jpeg", "ID_Combined.jpeg"]),
            "app":           _ls(cdir / "app" / "artifacts", exts=[".apk", ".aab"]),
            "payment_profile": [_pp(cdir)] if _pp(cdir) else [],
        },
        "tree":             _tree(cdir),
        "version":          _read_version(cn, name),
        "play_console":     _build_pc(cn, name, app_id, manifest, a_email, domain_val),
        "jobs":             jobs,
    })


@api_v2.route("/api/v2/company/<cn>/stage/<stage>", methods=["POST"])
def v2_run_stage(cn, stage):
    if stage not in STAGE_ORDER:
        return jsonify({"error": f"Unknown stage: {stage}"}), 400

    row = _find_row(cn)
    if not row:
        return jsonify({"error": "Company not found"}), 404

    if _job_get(cn, stage).get("status") == "running":
        return jsonify({"error": "Already running"}), 409

    _spawn(cn, stage, lambda: _exec_stage(cn, stage, _find_row(cn) or row))
    return jsonify({"status": "started", "cn": cn, "stage": stage})


@api_v2.route("/api/v2/company/<cn>/run-all", methods=["POST"])
def v2_run_all(cn):
    """Run all pending stages in order for one company."""
    row = _find_row(cn)
    if not row:
        return jsonify({"error": "Company not found"}), 404

    running = [s for s in STAGE_ORDER if _job_get(cn, s).get("status") == "running"]
    if running:
        return jsonify({"error": f"'{running[0]}' is already running"}), 409

    def _seq():
        with _co_lock(cn):
            for stage in STAGE_ORDER:
                fresh = _find_row(cn) or {}
                if _compute_stages(cn, fresh).get(stage):
                    continue  # already done
                key = _jk(cn, stage)
                _job_set(cn, stage, status="running",
                         started_at=datetime.now().isoformat(),
                         finished_at=None, error=None, log=None)
                _TEE.bind(key)
                try:
                    _exec_stage(cn, stage, _find_row(cn) or fresh)
                    log = _TEE.unbind()
                    _job_set(cn, stage, status="done",
                             finished_at=datetime.now().isoformat(),
                             log=log)
                except Exception as e:
                    log = _TEE.unbind()
                    _job_set(cn, stage, status="failed",
                             finished_at=datetime.now().isoformat(),
                             error=str(e),
                             log=log + "\n" + traceback.format_exc())
                    break  # stop on first failure

    threading.Thread(target=_seq, daemon=True).start()
    return jsonify({"status": "started"})


@api_v2.route("/api/v2/company/<cn>/job/<stage>")
def v2_job(cn, stage):
    j = _job_get(cn, stage)
    return jsonify(j if j else {"status": "idle"})


@api_v2.route("/api/v2/company/search")
def v2_search():
    q = request.args.get("q", "").strip()
    if len(q) < 2:
        return jsonify([])

    ql = q.lower()
    rows = _all_rows()
    results = []

    for row in rows:
        cn    = _cn8(row.get("Company Number"))
        name  = str(row.get("Company Name") or "")
        dirs  = str(row.get("Directors") or "")
        if ql in cn or ql in name.lower() or ql in dirs.lower():
            stages = _compute_stages(cn, row)
            results.append({
                "cn":        cn,
                "name":      name,
                "directors": dirs,
                "n_done":    sum(stages.values()),
                "stages":    stages,
                "source":    "local",
            })

    # Fall back to CH API if nothing found locally
    ch_error = None
    if not results:
        try:
            sys.path.insert(0, str(ROOT))
            from run_pipeline import ch_get
            data = ch_get(f"/search/companies?q={q}&items_per_page=8")
            for item in (data.get("items") or []):
                if item.get("company_status", "").lower() != "active": continue
                item_cn = _cn8(item.get("company_number", ""))
                if any(c["cn"] == item_cn for c in results): continue
                results.append({
                    "cn":        item_cn,
                    "name":      item.get("title", ""),
                    "directors": "",
                    "n_done":    0,
                    "stages":    {s: False for s in STAGE_ORDER},
                    "source":    "ch_api",
                })
        except Exception as e:
            ch_error = str(e)
            print(f"  CH search error: {e}")

    resp: dict = {"results": results[:10]}
    if ch_error:
        resp["ch_error"] = ch_error
    return jsonify(resp)


@api_v2.route("/api/v2/company/add", methods=["POST"])
def v2_add_company():
    data = request.get_json(force=True) or {}
    cn   = _cn8(data.get("cn") or data.get("company_number") or "")
    name = str(data.get("name") or data.get("company_name") or "").strip()
    if not cn or cn == "00000000":
        return jsonify({"error": "company_number required"}), 400

    existing = _find_row(cn)
    if existing:
        return jsonify({"status": "exists", "cn": cn, "name": str(existing.get("Company Name") or "")}), 200

    _add_company_row(cn, name)
    return jsonify({"status": "added", "cn": cn, "name": name}), 201


@api_v2.route("/api/v2/company/<cn>/version", methods=["POST"])
def v2_set_version(cn):
    """Save version_code + version_name for a company (does NOT rebuild — caller triggers that)."""
    row = _find_row(cn)
    if not row:
        return jsonify({"error": "Company not found"}), 404
    body = request.get_json(force=True, silent=True) or {}
    vc = body.get("version_code")
    vn = str(body.get("version_name", "")).strip()
    if not isinstance(vc, int) or vc < 1:
        return jsonify({"error": "version_code must be a positive integer"}), 400
    if not vn:
        return jsonify({"error": "version_name is required"}), 400
    name = str(row.get("Company Name") or "")
    _write_version(cn, name, vc, vn)
    return jsonify({"status": "saved", "version_code": vc, "version_name": vn})


@api_v2.route("/api/v2/company/<cn>/play-console", methods=["POST"])
def v2_save_play_console(cn):
    """Save all Play Console form sections for a company."""
    row = _find_row(cn)
    if not row:
        return jsonify({"error": "Company not found"}), 404
    body = request.get_json(force=True, silent=True) or {}
    name = str(row.get("Company Name") or "")

    app_name = str(body.get("app_name", "")).strip()
    if not app_name:
        return jsonify({"error": "app_name is required"}), 400
    if len(app_name) > 30:
        return jsonify({"error": "app_name must be ≤ 30 characters"}), 400
    pkg = str(body.get("package_name", "")).strip()
    if len(pkg) > 150:
        return jsonify({"error": "package_name must be ≤ 150 characters"}), 400
    short_desc = str(body.get("short_description", ""))
    if len(short_desc) > 80:
        return jsonify({"error": "short_description must be ≤ 80 characters"}), 400
    full_desc = str(body.get("full_description", ""))
    if len(full_desc) > 4000:
        return jsonify({"error": "full_description must be ≤ 4000 characters"}), 400

    _STR = [
        "app_name", "package_name", "default_language", "app_type", "app_access",
        "ratings_category", "ratings_email", "target_age", "financial_features",
        "health_features", "app_category", "contact_email", "contact_website",
        "short_description", "full_description",
        "release_name", "release_notes",
    ]
    _BOOL = [
        "is_free", "policy_confirmed", "signing_tos_accepted", "export_laws_accepted",
        "contains_ads", "q_violence", "q_sexual", "q_teen_rating", "q_profanity",
        "q_controlled_drugs", "q_gambling", "q_user_generated", "q_account_sharing",
        "q_location_sharing", "ratings_iarc_agreed", "legal_compliance",
        "data_collection", "families_policy", "is_government", "uses_advertising_id",
    ]
    data: dict = {}
    for f in _STR:
        data[f] = str(body.get(f, "")).strip()
    for f in _BOOL:
        data[f] = bool(body.get(f, False))

    _write_pc(cn, name, data)
    return jsonify({"status": "saved", **data})


@api_v2.route("/api/v2/company/<cn>/rerun-id", methods=["POST"])
def v2_rerun_id(cn):
    """Delete existing director ID files and regenerate a new one via Xbinder."""
    row = _find_row(cn)
    if not row:
        return jsonify({"error": "Company not found"}), 404
    if _job_get(cn, "director_id").get("status") == "running":
        return jsonify({"error": "Already running"}), 409
    _spawn(cn, "director_id", lambda: _exec_stage(cn, "director_id", _find_row(cn) or row))
    return jsonify({"status": "started", "cn": cn, "stage": "director_id"})


@api_v2.route("/api/v2/dl")
def v2_download():
    """Serve a file from pipeline_output (security: must be under DL_BASE).

    ?inline=1  — serve inline (for in-browser preview); default is attachment download.
    """
    path = request.args.get("path", "").strip()
    if not path:
        return abort(400)
    abs_path = Path(os.path.abspath(path))
    if not str(abs_path).startswith(str(DL_BASE.resolve())):
        return abort(403)
    if not abs_path.is_file():
        return abort(404)
    inline = request.args.get("inline", "0") == "1"
    return send_file(abs_path, as_attachment=not inline)


@api_v2.route("/api/v2/company/<cn>/archive", methods=["POST"])
def v2_archive_company(cn):
    """Archive a company (hide from main view, show in archive tab)."""
    row = _find_row(cn)
    if not row:
        return jsonify({"error": "Company not found"}), 404
    _update_excel_row(cn, {"Archived": "Yes"})
    return jsonify({"status": "archived", "cn": cn, "archived": True})


@api_v2.route("/api/v2/company/<cn>/unarchive", methods=["POST"])
def v2_unarchive_company(cn):
    """Unarchive a company (show in main view)."""
    row = _find_row(cn)
    if not row:
        return jsonify({"error": "Company not found"}), 404
    _update_excel_row(cn, {"Archived": "No"})
    return jsonify({"status": "unarchived", "cn": cn, "archived": False})


@api_v2.route("/api/v2/company/<cn>/status", methods=["POST"])
def v2_update_status(cn):
    """Update company status (Active, Declined, Pending, On Hold)."""
    row = _find_row(cn)
    if not row:
        return jsonify({"error": "Company not found"}), 404
    
    body = request.json or {}
    status = str(body.get("status", "")).strip()
    valid_statuses = ["Active", "Declined", "Pending", "On Hold"]
    
    if status and status not in valid_statuses:
        return jsonify({"error": f"Invalid status. Must be one of: {', '.join(valid_statuses)}"}), 400
    
    _update_excel_row(cn, {"Company Status": status})
    return jsonify({"status": "updated", "cn": cn, "company_status": status})


@api_v2.route("/api/v2/company/<cn>/notes", methods=["POST"])
def v2_update_notes(cn):
    """Update company notes (max 1000 characters)."""
    row = _find_row(cn)
    if not row:
        return jsonify({"error": "Company not found"}), 404
    
    body = request.json or {}
    notes = str(body.get("notes", "")).strip()
    
    if len(notes) > 1000:
        return jsonify({"error": "Notes must be 1000 characters or less"}), 400
    
    _update_excel_row(cn, {"Company Notes": notes})
    return jsonify({"status": "updated", "cn": cn, "notes": notes})


@api_v2.route("/health")
def v2_health():
    return jsonify({"status": "ok", "ts": datetime.now().isoformat()})
