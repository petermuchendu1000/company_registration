"""
Automate dev@ email forwarding on Namecheap for company domains.

For every company in the pipeline Excel that has a Domain assigned:
  1. Check if dev@{domain} forwarding already exists in Namecheap
  2. If not, create it — forwarding to the company's Assigned Email (Gmail)
  3. Update the company's manifest.json support_email to dev@{domain}

Run:
    python setup_dev_email.py [--dry-run]
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl
import requests

from check_ip import ensure_ip_whitelisted

BASE_DIR   = Path(__file__).parent
EXCEL_PATH = BASE_DIR / "pipeline_output" / "companies_pipeline.xlsx"
APPS_DIR   = BASE_DIR / "pipeline_output" / "apps"
API_URL    = "https://api.namecheap.com/xml.response"
NS         = {"nc": "http://api.namecheap.com/xml.response"}


def _api_base() -> dict:
    return {
        "ApiUser":  os.environ["NAMECHEAP_API_USER"],
        "ApiKey":   os.environ["NAMECHEAP_API_KEY"],
        "UserName": os.environ["NAMECHEAP_USERNAME"],
        "ClientIp": os.environ["NAMECHEAP_CLIENT_IP"],
    }


def get_existing_forwards(domain: str) -> list[tuple[str, str]]:
    """Return list of (mailbox, forward_to) tuples currently on the domain."""
    params = {**_api_base(), "Command": "namecheap.domains.dns.getEmailForwarding", "DomainName": domain}
    r = requests.get(API_URL, params=params, timeout=30)
    root = ET.fromstring(r.text)
    if root.get("Status") == "ERROR":
        errs = [e.text for e in root.findall(".//nc:Error", NS)]
        # External DNS — can't manage forwarding via API
        if any("DNS" in (e or "") for e in errs):
            return None  # type: ignore[return-value]
        raise RuntimeError(f"Namecheap error for {domain}: {errs}")
    forwards = []
    for fw in root.findall(".//nc:Forward", NS):
        mailbox = fw.get("mailbox", "")
        fwd_to  = fw.text or ""
        if mailbox and fwd_to:
            forwards.append((mailbox, fwd_to))
    return forwards


def set_email_forwarding(domain: str, forwards: list[tuple[str, str]]) -> None:
    """Replace all email forwarding rules for the domain."""
    params = {**_api_base(), "Command": "namecheap.domains.dns.setEmailForwarding", "DomainName": domain}
    for i, (mailbox, fwd_to) in enumerate(forwards, 1):
        params[f"mailbox{i}"]   = mailbox
        params[f"ForwardTo{i}"] = fwd_to
    r = requests.get(API_URL, params=params, timeout=30)
    root = ET.fromstring(r.text)
    if root.get("Status") == "ERROR":
        errs = [e.text for e in root.findall(".//nc:Error", NS)]
        raise RuntimeError(f"Failed to set forwarding on {domain}: {errs}")


def update_manifest_email(company_number: str, new_email: str) -> bool:
    manifest_path = APPS_DIR / company_number / "manifest.json"
    if not manifest_path.exists():
        return False
    data = json.loads(manifest_path.read_text(encoding="utf-8"))
    if data.get("support_email") == new_email:
        return False  # already correct
    data["support_email"] = new_email
    manifest_path.write_text(json.dumps(data, indent=2), encoding="utf-8")
    return True


def load_company_domains(wb) -> list[tuple[str, str, str]]:
    """Return (company_number, domain, assigned_email) for rows with a domain set."""
    ws = wb.active
    header = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]

    def col(name: str) -> int | None:
        for i, h in enumerate(header, 1):
            if h == name.lower():
                return i
        return None

    cn_col     = col("company number")
    domain_col = col("domain")
    email_col  = col("assigned email")

    if not all([cn_col, domain_col, email_col]):
        raise RuntimeError(f"Missing columns. Found: {header}")

    rows = []
    for r in range(2, ws.max_row + 1):
        cn     = str(ws.cell(r, cn_col).value or "").strip()
        domain = str(ws.cell(r, domain_col).value or "").strip()
        email  = str(ws.cell(r, email_col).value or "").strip()
        if cn and domain and email:
            rows.append((cn, domain, email))
    return rows


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    ensure_ip_whitelisted(verbose=True)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    companies = load_company_domains(wb)
    wb.close()

    print(f"\nFound {len(companies)} companies with domains assigned.\n")

    for cn, domain, assigned_email in companies:
        dev_email = f"dev@{domain}"
        print(f"[{cn}] {domain}")

        existing = get_existing_forwards(domain)
        if existing is None:
            print(f"  SKIP — external DNS, can't manage via API")
            continue

        already = any(mb == "dev" for mb, _ in existing)
        if already:
            print(f"  OK   — {dev_email} already set up")
        else:
            new_forwards = existing + [("dev", assigned_email)]
            if args.dry_run:
                print(f"  DRY  — would create {dev_email} -> {assigned_email}")
            else:
                set_email_forwarding(domain, new_forwards)
                print(f"  SET  — {dev_email} -> {assigned_email}")

        # Update manifest regardless of dry-run flag so we can see what would change
        manifest_path = APPS_DIR / cn / "manifest.json"
        if manifest_path.exists():
            data = json.loads(manifest_path.read_text(encoding="utf-8"))
            current = data.get("support_email", "")
            if current != dev_email:
                if args.dry_run:
                    print(f"  DRY  — would update manifest support_email: {current!r} -> {dev_email!r}")
                else:
                    data["support_email"] = dev_email
                    manifest_path.write_text(json.dumps(data, indent=2), encoding="utf-8")
                    print(f"  UPD  — manifest support_email updated to {dev_email}")
            else:
                print(f"  OK   — manifest already has {dev_email}")
        else:
            print(f"  --   — no manifest yet (app not yet generated)")

        print()


if __name__ == "__main__":
    main()
