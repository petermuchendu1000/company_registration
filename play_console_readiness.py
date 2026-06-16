"""Build Google Play Console readiness records without touching Google.

This module turns the pipeline's company/domain/email data into a structured
"dossier" a human can use while creating a Play Console organization account.
It deliberately does not automate Play Console or Search Console.
"""

from __future__ import annotations

import json
import os
import re
from datetime import datetime
from pathlib import Path


BASE_DIR = os.path.dirname(__file__)
OUTPUT_DIR = os.path.join(BASE_DIR, "pipeline_output")
DOSSIER_DIR = os.path.join(OUTPUT_DIR, "play_console_dossiers")
PHONE_FILE_CANDIDATES = [
    os.path.join(BASE_DIR, "physical_numbers.json"),
    os.path.join(BASE_DIR, "phone_numbers.json"),
]
PHONE_XLSX_CANDIDATES = [
    os.path.join(BASE_DIR, "physical_numbers.xlsx"),
    os.path.join(BASE_DIR, "phone_numbers.xlsx"),
    os.path.join(BASE_DIR, "phones.xlsx"),
]


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _status(value: bool) -> str:
    return "ready" if value else "missing"


def _split_director_name(raw_name: str) -> dict[str, str]:
    """Parse Companies House officer names such as 'SURNAME, First Middle'."""
    name = _clean(raw_name)
    if not name:
        return {"first_name": "", "last_name": "", "display_name": ""}

    if "," in name:
        last, rest = name.split(",", 1)
        first_parts = [p for p in rest.strip().split() if p]
        first = first_parts[0].title() if first_parts else ""
        last = last.strip().title()
    else:
        parts = [p for p in name.split() if p]
        first = parts[0].title() if parts else ""
        last = parts[-1].title() if len(parts) > 1 else ""

    display = " ".join(p for p in [first, last] if p).strip() or name.title()
    return {"first_name": first, "last_name": last, "display_name": display}


def _first_director(details: dict) -> dict[str, str]:
    directors = details.get("directors") or []
    if directors and isinstance(directors, list):
        first = directors[0] or {}
        parsed = _split_director_name(first.get("name", ""))
        return {
            "name": first.get("name", ""),
            "display_name": parsed["display_name"],
            "first_name": parsed["first_name"],
            "last_name": parsed["last_name"],
            "nationality": first.get("nationality", ""),
            "appointed_on": first.get("appointed_on", ""),
            "role": "Director",
            "source": "Companies House officers API",
        }

    names = [n.strip() for n in _clean(details.get("director_names")).split(";") if n.strip()]
    raw = names[0] if names else ""
    parsed = _split_director_name(raw)
    nationalities = [n.strip() for n in _clean(details.get("director_nationalities")).split(";") if n.strip()]
    return {
        "name": raw,
        "display_name": parsed["display_name"],
        "first_name": parsed["first_name"],
        "last_name": parsed["last_name"],
        "nationality": nationalities[0] if nationalities else "",
        "appointed_on": "",
        "role": "Director",
        "source": "Companies Pipeline Excel",
    }


def developer_email_for_domain(domain: str) -> str:
    domain = _clean(domain).lower()
    return f"dev@{domain}" if domain else ""


def _load_phone_assignments() -> dict[str, str]:
    """Optional local phone assignment map.

    Supported JSON shapes:
      {"13510663": "+441234567890"}
      {"13510663": {"phone": "+441234567890"}}

    The user said physical numbers already exist; this lets the pipeline use
    them once added locally without requiring a Google interaction.
    """
    for path in PHONE_FILE_CANDIDATES:
        if not os.path.exists(path):
            continue
        try:
            with open(path, "r", encoding="utf-8") as f:
                raw = json.load(f)
        except Exception:
            continue

        out: dict[str, str] = {}
        if isinstance(raw, dict):
            for key, value in raw.items():
                if isinstance(value, dict):
                    phone = value.get("phone") or value.get("number") or value.get("phone_number")
                else:
                    phone = value
                if phone:
                    out[str(key).strip()] = str(phone).strip()
        return out
    for path in PHONE_XLSX_CANDIDATES:
        if not os.path.exists(path):
            continue
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            company_idx = _first_header_index(headers, ["company number", "company_number", "company"])
            phone_idx = _first_header_index(headers, ["phone", "phone number", "phone_number", "number"])
            email_idx = _first_header_index(headers, ["gmail", "email", "assigned email"])
            out = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                phone = row[phone_idx] if phone_idx is not None and len(row) > phone_idx else ""
                if not phone:
                    continue
                company = row[company_idx] if company_idx is not None and len(row) > company_idx else ""
                email = row[email_idx] if email_idx is not None and len(row) > email_idx else ""
                key = str(company or email or "").strip()
                if key:
                    out[key] = str(phone).strip()
            wb.close()
            return out
        except Exception:
            continue
    return {}


def _first_header_index(headers: list[str], names: list[str]) -> int | None:
    for name in names:
        if name in headers:
            return headers.index(name)
    return None


def _phone_for_company(company_number: str, result: dict) -> str:
    explicit = (
        result.get("phone", {}).get("phone_number")
        if isinstance(result.get("phone"), dict)
        else ""
    )
    if explicit:
        return _clean(explicit)
    assignments = _load_phone_assignments()
    email = _clean((result.get("email") or {}).get("email"))
    return assignments.get(_clean(company_number), "") or assignments.get(email, "")


def build_readiness_record(result: dict) -> dict:
    """Create a complete Play Console readiness record for one company."""
    company_number = _clean(result.get("company_number"))
    details = result.get("details") or {}
    duns = result.get("duns") or {}
    certificate = result.get("certificate") or {}
    domain_info = result.get("domain") or {}
    email_info = result.get("email") or {}
    play = result.get("play_console") or {}

    domain = _clean(domain_info.get("domain"))
    account_email = _clean(email_info.get("email"))
    developer_email = _clean(play.get("developer_email")) or developer_email_for_domain(domain)
    organization_phone = _phone_for_company(company_number, result)
    director = _first_director(details)
    certificate_path = _clean(certificate.get("path"))
    certificate_exists = bool(certificate_path and os.path.exists(certificate_path))

    dnb_name = _clean(duns.get("dnb_name") or duns.get("company_name"))
    dnb_address = _clean(duns.get("dnb_address"))
    ch_name = _clean(details.get("company_name"))
    ch_address = _clean(details.get("address"))
    dnb_name_match = bool(dnb_name and ch_name and _normalize_name(dnb_name) == _normalize_name(ch_name))

    forwarding = play.get("developer_email_forwarding") or {}
    txt = play.get("google_txt") or {}

    record = {
        "company_number": company_number,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "account_owner": {
            "gmail_login": account_email,
            "account_name": _clean(email_info.get("account_name")),
            "source": "emails.xlsx",
            "phone": organization_phone,
            "notes": "Human must sign in manually. Do not automate Google.",
        },
        "organization": {
            "account_type": "Organization",
            "legal_name": ch_name,
            "company_number": company_number,
            "company_status": _clean(details.get("company_status")),
            "company_type": _clean(details.get("company_type")),
            "date_of_creation": _clean(details.get("date_of_creation")),
            "sic_codes": _clean(details.get("sic_codes")),
            "legal_address": ch_address,
            "organization_phone": organization_phone,
            "organization_size": "1-10",
            "organization_category": "Company/business",
            "source": "Companies House",
        },
        "authorized_representative": {
            **director,
            "verification_basis": "Use this active director as the authorized representative.",
            "id_document_status": "manual_required",
            "id_quality_checklist": [
                "Valid and not expired",
                "Color image",
                "Clear and well lit",
                "Not a photocopy",
                "Name matches the director/representative",
            ],
        },
        "duns": {
            "duns_number": _clean(duns.get("duns_number")),
            "status": _clean(duns.get("status")),
            "dnb_legal_name": dnb_name,
            "dnb_address": dnb_address,
            "dnb_name_matches_companies_house": dnb_name_match,
            "dnb_address_match_status": "unknown" if not dnb_address else "needs_review",
        },
        "domain": {
            "domain": domain,
            "registration_status": _clean(domain_info.get("status")),
            "registration_cost": _clean(domain_info.get("charged")),
            "namecheap_account": "existing configured Namecheap API account",
            "registration_contacts_source": _clean(domain_info.get("contacts_source")) or "pipeline/default",
        },
        "developer_contact": {
            "private_contact_email": account_email,
            "public_developer_email": developer_email,
            "developer_phone": organization_phone,
            "app_support_email": developer_email,
            "email_forwarding": {
                "mailbox": "dev",
                "from": developer_email,
                "to": account_email,
                "status": _clean(forwarding.get("status")) or "not_configured",
                "provider": "Namecheap email forwarding",
                "can_receive": bool(developer_email and account_email),
                "can_send": False,
                "note": "Namecheap forwarding receives mail only; use Private Email/SMTP if sending as dev@domain is required.",
            },
        },
        "dns_verification": {
            "google_txt_host": _clean(txt.get("hostname")) or "@",
            "google_txt_value": _clean(txt.get("value")) or "PENDING_HUMAN_GOOGLE_TOKEN",
            "status": _clean(txt.get("status")) or "token_pending",
            "provider": "Namecheap DNS setHosts",
            "note": "Human copies TXT token from Google; pipeline can apply it to Namecheap DNS only.",
        },
        "documents": {
            "certificate_path": certificate_path,
            "certificate_exists": certificate_exists,
            "organization_document_status": "ready" if certificate_exists else "missing",
            "representative_id_status": "manual_required",
        },
        "payments": {
            "registration_fee": "USD 25 one-time",
            "payment_method_status": "manual_required",
            "billing_profile_note": "Google Payments profile legal name/address should match D&B and Companies House as closely as possible.",
        },
    }
    record["readiness"] = _readiness_summary(record)
    return record


def _normalize_name(name: str) -> str:
    name = name.upper()
    name = re.sub(r"\b(LTD|LIMITED|PLC|LLP|INC|CORP|LLC)\b\.?", "", name)
    name = re.sub(r"[^A-Z0-9]+", "", name)
    return name


def _readiness_summary(record: dict) -> dict:
    checks = {
        "gmail_owner_available": bool(record["account_owner"]["gmail_login"]),
        "company_data_available": bool(record["organization"]["legal_name"] and record["organization"]["legal_address"]),
        "director_available": bool(record["authorized_representative"]["name"]),
        "duns_available": bool(record["duns"]["duns_number"]),
        "domain_available": bool(record["domain"]["domain"]),
        "developer_email_defined": bool(record["developer_contact"]["public_developer_email"]),
        "developer_email_forwarding_configured": record["developer_contact"]["email_forwarding"]["status"] == "configured",
        "organization_phone_available": bool(record["organization"]["organization_phone"]),
        "certificate_available": bool(record["documents"]["certificate_exists"]),
        "google_txt_token_added": record["dns_verification"]["status"] == "configured",
        "representative_id_ready": record["documents"]["representative_id_status"] == "ready",
        "payment_method_ready": record["payments"]["payment_method_status"] == "ready",
    }
    blocking_missing = [name for name, ok in checks.items() if not ok]
    return {
        "checks": checks,
        "ready_for_human_signup": not blocking_missing,
        "blocking_missing": blocking_missing,
    }


def write_readiness_sheet(workbook, results: list[dict]) -> None:
    """Add a Play Console Readiness sheet to an openpyxl workbook."""
    from openpyxl.styles import Alignment, Font, PatternFill

    if "Play Console Readiness" in workbook.sheetnames:
        del workbook["Play Console Readiness"]
    ws = workbook.create_sheet("Play Console Readiness")
    headers = [
        "No.", "Company Number", "Organization Legal Name", "Account Type",
        "Gmail Owner", "Authorized Representative", "Representative Role",
        "DUNS Number", "D&B Name", "D&B Name Match", "Organization Phone",
        "Domain", "Developer Email", "Developer Email Forwarding",
        "Google TXT Status", "Certificate", "Representative ID",
        "Payment Method", "Missing Before Signup",
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color="23422F", end_color="23422F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, result in enumerate(results, 1):
        rec = build_readiness_record(result)
        missing = ", ".join(rec["readiness"]["blocking_missing"])
        ws.append([
            i,
            rec["company_number"],
            rec["organization"]["legal_name"],
            rec["organization"]["account_type"],
            rec["account_owner"]["gmail_login"],
            rec["authorized_representative"]["display_name"] or rec["authorized_representative"]["name"],
            rec["authorized_representative"]["role"],
            rec["duns"]["duns_number"],
            rec["duns"]["dnb_legal_name"],
            "Yes" if rec["duns"]["dnb_name_matches_companies_house"] else "Review",
            rec["organization"]["organization_phone"],
            rec["domain"]["domain"],
            rec["developer_contact"]["public_developer_email"],
            rec["developer_contact"]["email_forwarding"]["status"],
            rec["dns_verification"]["status"],
            rec["documents"]["organization_document_status"],
            rec["documents"]["representative_id_status"],
            rec["payments"]["payment_method_status"],
            missing,
        ])

    for col in ws.columns:
        col_letter = col[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)
    ws.freeze_panes = "A2"


def export_dossier(results: list[dict], output_dir: str = DOSSIER_DIR) -> dict[str, str]:
    """Write Markdown and JSON dossiers for all results."""
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    records = [build_readiness_record(r) for r in results]
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = os.path.join(output_dir, f"play_console_readiness_{stamp}.json")
    md_path = os.path.join(output_dir, f"play_console_dossier_{stamp}.md")

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, ensure_ascii=False)

    with open(md_path, "w", encoding="utf-8") as f:
        f.write("# Google Play Console Organization Signup Dossier\n\n")
        f.write("This dossier is for human form filling only. Do not automate Google Play Console or Search Console.\n\n")
        for rec in records:
            _write_record_markdown(f, rec)
    return {"json": json_path, "markdown": md_path, "count": str(len(records))}


def _write_record_markdown(f, rec: dict) -> None:
    f.write(f"## {rec['organization']['legal_name']} ({rec['company_number']})\n\n")
    f.write(f"- Overall ready for human signup: {rec['readiness']['ready_for_human_signup']}\n")
    if rec["readiness"]["blocking_missing"]:
        f.write(f"- Missing before signup: {', '.join(rec['readiness']['blocking_missing'])}\n")
    f.write("\n")

    sections = [
        ("Account Owner", rec["account_owner"]),
        ("Organization", rec["organization"]),
        ("Authorized Representative", rec["authorized_representative"]),
        ("DUNS", rec["duns"]),
        ("Domain", rec["domain"]),
        ("Developer Contact", rec["developer_contact"]),
        ("DNS Verification", rec["dns_verification"]),
        ("Documents", rec["documents"]),
        ("Payments", rec["payments"]),
    ]
    for title, data in sections:
        f.write(f"### {title}\n\n")
        _write_dict(f, data)
        f.write("\n")


def _write_dict(f, data: dict, prefix: str = "") -> None:
    for key, value in data.items():
        label = key.replace("_", " ").title()
        if isinstance(value, dict):
            f.write(f"- {prefix}{label}:\n")
            _write_dict(f, value, prefix="  ")
        elif isinstance(value, list):
            f.write(f"- {prefix}{label}: {', '.join(str(v) for v in value)}\n")
        else:
            f.write(f"- {prefix}{label}: {value}\n")
