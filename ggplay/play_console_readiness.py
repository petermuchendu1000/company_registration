"""Build Google Play Console readiness records without touching Google.

This module turns the pipeline's company/domain/email data into a structured
"dossier" a human can use while creating a Play Console organization account.
It deliberately does not automate Play Console or Search Console.
"""

from __future__ import annotations

import json
import os
import re
import socket
from datetime import datetime
from pathlib import Path


BASE_DIR = os.path.dirname(__file__)
OUTPUT_DIR = os.path.join(BASE_DIR, "pipeline_output")
DOSSIER_DIR = os.path.join(OUTPUT_DIR, "play_console_dossiers")
PHONE_FILE_CANDIDATES = [
    os.path.join(BASE_DIR, "physical_numbers.json"),
    os.path.join(BASE_DIR, "phone_numbers.json"),
]
PHONE_XLSX_CANDIDATES = [
    os.path.join(BASE_DIR, "physical_numbers.xlsx"),
    os.path.join(BASE_DIR, "phone_numbers.xlsx"),
    os.path.join(BASE_DIR, "phones.xlsx"),
]


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _status(value: bool) -> str:
    return "ready" if value else "missing"


def _split_director_name(raw_name: str) -> dict[str, str]:
    """Parse Companies House officer names such as 'SURNAME, First Middle'."""
    name = _clean(raw_name)
    if not name:
        return {"first_name": "", "last_name": "", "display_name": ""}

    if "," in name:
        last, rest = name.split(",", 1)
        first_parts = [p for p in rest.strip().split() if p]
        first = first_parts[0].title() if first_parts else ""
        last = last.strip().title()
    else:
        parts = [p for p in name.split() if p]
        first = parts[0].title() if parts else ""
        last = parts[-1].title() if len(parts) > 1 else ""

    display = " ".join(p for p in [first, last] if p).strip() or name.title()
    return {"first_name": first, "last_name": last, "display_name": display}


def _first_director(details: dict) -> dict[str, str]:
    directors = details.get("directors") or []
    if directors and isinstance(directors, list):
        first = directors[0] or {}
        parsed = _split_director_name(first.get("name", ""))
        return {
            "name": first.get("name", ""),
            "display_name": parsed["display_name"],
            "first_name": parsed["first_name"],
            "last_name": parsed["last_name"],
            "nationality": first.get("nationality", ""),
            "appointed_on": first.get("appointed_on", ""),
            "role": "Director",
            "source": "Companies House officers API",
        }

    names = [n.strip() for n in _clean(details.get("director_names")).split(";") if n.strip()]
    raw = names[0] if names else ""
    parsed = _split_director_name(raw)
    nationalities = [n.strip() for n in _clean(details.get("director_nationalities")).split(";") if n.strip()]
    return {
        "name": raw,
        "display_name": parsed["display_name"],
        "first_name": parsed["first_name"],
        "last_name": parsed["last_name"],
        "nationality": nationalities[0] if nationalities else "",
        "appointed_on": "",
        "role": "Director",
        "source": "Companies Pipeline Excel",
    }


def developer_email_for_domain(domain: str) -> str:
    domain = _clean(domain).lower()
    return f"dev@{domain}" if domain else ""


def _load_phone_assignments() -> dict[str, str]:
    """Optional local phone assignment map.

    Supported JSON shapes:
      {"13510663": "+441234567890"}
      {"13510663": {"phone": "+441234567890"}}

    The user said physical numbers already exist; this lets the pipeline use
    them once added locally without requiring a Google interaction.
    """
    for path in PHONE_FILE_CANDIDATES:
        if not os.path.exists(path):
            continue
        try:
            with open(path, "r", encoding="utf-8") as f:
                raw = json.load(f)
        except Exception:
            continue

        out: dict[str, str] = {}
        if isinstance(raw, dict):
            for key, value in raw.items():
                if isinstance(value, dict):
                    phone = value.get("phone") or value.get("number") or value.get("phone_number")
                else:
                    phone = value
                if phone:
                    out[str(key).strip()] = str(phone).strip()
        return out
    for path in PHONE_XLSX_CANDIDATES:
        if not os.path.exists(path):
            continue
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            company_idx = _first_header_index(headers, ["company number", "company_number", "company"])
            phone_idx = _first_header_index(headers, ["phone", "phone number", "phone_number", "number"])
            email_idx = _first_header_index(headers, ["gmail", "email", "assigned email"])
            out = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                phone = row[phone_idx] if phone_idx is not None and len(row) > phone_idx else ""
                if not phone:
                    continue
                company = row[company_idx] if company_idx is not None and len(row) > company_idx else ""
                email = row[email_idx] if email_idx is not None and len(row) > email_idx else ""
                key = str(company or email or "").strip()
                if key:
                    out[key] = str(phone).strip()
            wb.close()
            return out
        except Exception:
            continue
    return {}


def _first_header_index(headers: list[str], names: list[str]) -> int | None:
    for name in names:
        if name in headers:
            return headers.index(name)
    return None


def _phone_for_company(company_number: str, result: dict) -> dict:
    """Return a full phone record for the company via phone_pool."""
    # Honour an explicitly pre-set phone on the result dict first
    explicit = (
        result.get("phone", {}).get("phone_number")
        if isinstance(result.get("phone"), dict)
        else ""
    )
    if explicit:
        phone = _clean(explicit)
        return {
            "organization_phone": phone,
            "contact_phone": phone,
            "developer_phone": phone,
            "phone_type": "sim",
            "otp_capable": True,
            "source": "result.phone",
            "is_dummy": False,
        }

    # Fall back to phone_pool (assigns a dummy +254 if no real number yet)
    try:
        from phone_pool import assign_phone
        return assign_phone(_clean(company_number))
    except Exception:
        pass

    # Last resort: legacy JSON/XLSX lookup
    assignments = _load_phone_assignments()
    email = _clean((result.get("email") or {}).get("email"))
    phone = assignments.get(_clean(company_number), "") or assignments.get(email, "")
    if phone:
        return {
            "organization_phone": phone,
            "contact_phone": phone,
            "developer_phone": phone,
            "phone_type": "unknown",
            "otp_capable": True,
            "source": "legacy_json",
            "is_dummy": False,
        }
    return {}


def _normalize_address(addr: str) -> str:
    """Normalise an address string for comparison.

    Collapses both full-form and abbreviated variants to the same abbreviation
    so "High Street" and "High St" produce identical normalised strings.
    """
    a = addr.upper()
    # Collapse full forms → abbreviations (matching both directions)
    a = re.sub(r"\bSTREET\b", "ST", a)       # Street  → St
    a = re.sub(r"\bSAINT\b",  "ST", a)       # Saint   → St  (place names)
    a = re.sub(r"\bROAD\b",   "RD", a)       # Road    → Rd
    a = re.sub(r"\bAVENUE\b", "AVE", a)      # Avenue  → Ave
    a = re.sub(r"\bDRIVE\b",  "DR", a)       # Drive   → Dr
    a = re.sub(r"\bCLOSE\b",  "CL", a)       # Close   → Cl
    a = re.sub(r"\bCOURT\b",  "CT", a)       # Court   → Ct
    a = re.sub(r"\bLANE\b",   "LN", a)       # Lane    → Ln
    a = re.sub(r"\bPLACE\b",  "PL", a)       # Place   → Pl
    a = re.sub(r"\bGARDENS\b","GDNS", a)     # Gardens → Gdns
    a = re.sub(r"\bCRESCENT\b","CRES", a)   # Crescent → Cres
    # Strip country suffixes that differ between CH and D&B
    a = re.sub(r",?\s*(ENGLAND|WALES|SCOTLAND|UNITED KINGDOM|UK|GREAT BRITAIN|GB)\s*$", "", a)
    # Keep only alphanumerics
    a = re.sub(r"[^A-Z0-9]+", "", a)
    return a


def _compare_names(ch: str, dnb: str) -> dict:
    """Strict name comparison between Companies House and D&B."""
    if not ch or not dnb:
        return {"match": False, "detail": "one or both names missing", "normalized_ch": ch, "normalized_dnb": dnb}
    n_ch = _normalize_name(ch)
    n_dnb = _normalize_name(dnb)
    if n_ch == n_dnb:
        return {"match": True, "detail": "exact_match", "normalized_ch": n_ch, "normalized_dnb": n_dnb}
    # Partial: one contains the other
    if n_ch in n_dnb or n_dnb in n_ch:
        return {"match": False, "detail": f"partial_match: '{ch}' vs '{dnb}'", "normalized_ch": n_ch, "normalized_dnb": n_dnb}
    return {"match": False, "detail": f"mismatch: '{ch}' vs '{dnb}'", "normalized_ch": n_ch, "normalized_dnb": n_dnb}


def _compare_addresses(ch: str, dnb: str) -> dict:
    """Strict address comparison between Companies House and D&B."""
    if not ch or not dnb:
        return {"match": False, "detail": "one or both addresses missing"}
    n_ch = _normalize_address(ch)
    n_dnb = _normalize_address(dnb)
    if n_ch == n_dnb:
        return {"match": True, "detail": "exact_match"}
    if n_ch in n_dnb or n_dnb in n_ch:
        return {"match": False, "detail": f"partial_match: abbreviated vs full — '{ch}' vs '{dnb}'"}
    return {"match": False, "detail": f"mismatch: '{ch}' vs '{dnb}'"}


def _developer_display_name(ch_name: str) -> str:
    """Generate a public-facing developer display name from the legal name."""
    name = ch_name.strip()
    # Remove legal suffixes
    name = re.sub(r"\b(LIMITED|LTD|PLC|LLP|CIC|CIO|COMMUNITY INTEREST COMPANY)\b\.?", "", name, flags=re.IGNORECASE)
    name = name.strip(" ,.").strip()
    # Title-case
    name = name.title()
    return name or ch_name.title()


def _check_txt_propagated(domain: str, expected_token: str) -> bool:
    """Check if the Google Search Console TXT record has propagated via DNS."""
    if not domain or not expected_token or expected_token.startswith("PENDING"):
        return False
    try:
        import subprocess
        result = subprocess.run(
            ["nslookup", "-type=TXT", domain],
            capture_output=True, text=True, timeout=10
        )
        return expected_token in (result.stdout or "")
    except Exception:
        pass
    # Fallback: socket DNS (no TXT support in standard library — just return False)
    return False


def build_readiness_record(result: dict) -> dict:
    """Create a complete Play Console readiness record for one company."""
    company_number = _clean(result.get("company_number"))
    details = result.get("details") or {}
    duns = result.get("duns") or {}
    certificate = result.get("certificate") or {}
    domain_info = result.get("domain") or {}
    email_info = result.get("email") or {}
    play = result.get("play_console") or {}
    rep_id_info = result.get("representative_id") or {}
    payments_info = result.get("payments") or {}
    google_acct = result.get("google_account") or {}

    domain = _clean(domain_info.get("domain"))
    account_email = _clean(email_info.get("email"))
    developer_email = _clean(play.get("developer_email")) or developer_email_for_domain(domain)
    organization_website = f"https://{domain}" if domain else ""

    # Phone pool
    phone_record = _phone_for_company(company_number, result)
    organization_phone = phone_record.get("organization_phone", "")

    director = _first_director(details)
    # Contact name derived from director
    contact_name = director.get("display_name") or director.get("name", "")

    certificate_path = _clean(certificate.get("path"))
    certificate_exists = bool(certificate_path and os.path.exists(certificate_path))

    dnb_name = _clean(duns.get("dnb_name") or duns.get("company_name"))
    dnb_address = _clean(duns.get("dnb_address"))
    ch_name = _clean(details.get("company_name"))
    ch_address = _clean(details.get("address"))

    name_cmp = _compare_names(ch_name, dnb_name)
    addr_cmp = _compare_addresses(ch_address, dnb_address)
    cert_name_match = _compare_names(ch_name, dnb_name)["match"]  # cert shows CH name; must equal D&B

    forwarding = play.get("developer_email_forwarding") or {}
    txt = play.get("google_txt") or {}
    txt_token = _clean(txt.get("value") or "")
    txt_propagated = _check_txt_propagated(domain, txt_token) if txt_token else False
    txt_status = "verified" if txt_propagated else (_clean(txt.get("status")) or "token_pending")

    # Developer display name: use the Google account name (from Gmail) as the primary
    # source — this is what Google pre-populates when the user signs into Play Console.
    # Fall back to a cleaned company name only when no Gmail account name is available.
    account_name = _clean(email_info.get("account_name"))
    developer_display_name = account_name or (_developer_display_name(ch_name) if ch_name else "")

    record = {
        "company_number": company_number,
        "generated_at": datetime.now().isoformat(timespec="seconds"),

        "account_owner": {
            "gmail_login": account_email,
            "account_name": _clean(email_info.get("account_name")),
            "source": "emails.xlsx",
            "2sv_enabled": bool(google_acct.get("2sv_enabled", False)),
            "recovery_email": _clean(google_acct.get("recovery_email", "")),
            "recovery_phone": _clean(google_acct.get("recovery_phone", "")),
            "contact_email_otp_verified": bool(google_acct.get("contact_email_otp_verified", False)),
            "contact_phone_otp_verified": bool(google_acct.get("contact_phone_otp_verified", False)),
            "developer_email_otp_verified": bool(google_acct.get("developer_email_otp_verified", False)),
            "developer_phone_otp_verified": bool(google_acct.get("developer_phone_otp_verified", False)),
            "notes": "Human must sign in manually. Do not automate Google.",
        },

        "organization": {
            "account_type": "Organization",
            "legal_name": ch_name,
            "developer_display_name": developer_display_name,
            "organization_website": organization_website,
            "company_number": company_number,
            "company_status": _clean(details.get("company_status")),
            "company_type": _clean(details.get("company_type")),
            "date_of_creation": _clean(details.get("date_of_creation")),
            "sic_codes": _clean(details.get("sic_codes")),
            "legal_address": ch_address,
            "organization_phone": organization_phone,
            "organization_size": "1-10",
            "organization_category": "Company/business",
            "source": "Companies House",
        },

        "contact": {
            "contact_name": contact_name,
            "contact_email": account_email,
            "contact_phone": phone_record.get("contact_phone", organization_phone),
            "note": "Private contact info shown only to Google. Both email and phone must receive OTP during signup.",
        },

        "authorized_representative": {
            **director,
            "full_name_raw": director.get("name", ""),
            "full_name_as_on_id": rep_id_info.get("full_name_as_on_id", ""),
            "personal_address": rep_id_info.get("personal_address", ""),
            "verification_basis": "Use this active director as the authorized representative.",
            "id_type": rep_id_info.get("id_type", ""),
            "id_country": rep_id_info.get("id_country", ""),
            "id_expiry_date": rep_id_info.get("id_expiry_date", ""),
            "id_image_front": rep_id_info.get("id_image_front", ""),
            "id_image_back": rep_id_info.get("id_image_back", ""),
            "id_quality_check": rep_id_info.get("id_quality_check", "pending"),
            "id_document_status": "ready" if rep_id_info.get("id_image_front") else "manual_required",
            "proof_of_address_type": rep_id_info.get("proof_of_address_type", ""),
            "proof_of_address_path": rep_id_info.get("proof_of_address_path", ""),
            "proof_of_address_status": "ready" if rep_id_info.get("proof_of_address_path") else "manual_required",
            "id_quality_checklist": [
                "Valid and not expired",
                "Color image",
                "Clear and well lit",
                "Not a photocopy",
                "Name matches the director/representative exactly",
                "UK/EEA: include separate proof-of-address document",
            ],
        },

        "duns": {
            "duns_number": _clean(duns.get("duns_number")),
            "status": _clean(duns.get("status")),
            "dnb_legal_name": dnb_name,
            "dnb_address": dnb_address,
            "ch_name": ch_name,
            "ch_address": ch_address,
            "name_match": name_cmp["match"],
            "name_match_detail": name_cmp["detail"],
            "address_match": addr_cmp["match"],
            "address_match_detail": addr_cmp["detail"],
            "certificate_name_matches_duns": cert_name_match,
        },

        "domain": {
            "domain": domain,
            "registration_status": _clean(domain_info.get("status")),
            "registration_cost": _clean(domain_info.get("charged")),
            "namecheap_account": "existing configured Namecheap API account",
            "registration_contacts_source": _clean(domain_info.get("contacts_source")) or "pipeline/default",
        },

        "developer_contact": {
            "private_contact_email": account_email,
            "public_developer_email": developer_email,
            "developer_phone": phone_record.get("developer_phone", organization_phone),
            "app_support_email": developer_email,
            "phone_is_dummy": phone_record.get("is_dummy", True),
            "phone_otp_capable": phone_record.get("otp_capable", False),
            "email_forwarding": {
                "mailbox": "dev",
                "from": developer_email,
                "to": account_email,
                "status": _clean(forwarding.get("status")) or "not_configured",
                "provider": "Namecheap email forwarding",
                "can_receive": bool(developer_email and account_email),
                "can_send": False,
                "note": "Namecheap forwarding receives mail only; use Private Email/SMTP if sending as dev@domain is required.",
            },
        },

        "dns_verification": {
            "google_txt_host": _clean(txt.get("hostname")) or "@",
            "google_txt_value": txt_token or "PENDING_HUMAN_GOOGLE_TOKEN",
            "status": txt_status,
            "propagated": txt_propagated,
            "provider": "Namecheap DNS setHosts",
            "note": "Human copies TXT token from Google Search Console; pipeline applies it to Namecheap DNS. No website needed — TXT record proves domain ownership.",
        },

        "documents": {
            "certificate_path": certificate_path,
            "certificate_exists": certificate_exists,
            "organization_document_status": "ready" if certificate_exists else "missing",
            "representative_id_status": "ready" if rep_id_info.get("id_image_front") else "manual_required",
            "proof_of_address_status": "ready" if rep_id_info.get("proof_of_address_path") else "manual_required",
        },

        "payments": {
            "registration_fee": "USD 25 one-time",
            "card_last4": _clean(payments_info.get("card_last4", "")),
            "card_type": _clean(payments_info.get("card_type", "")),
            "card_holder_name": _clean(payments_info.get("card_holder_name", "")),
            "card_linked_to_payments_profile": bool(payments_info.get("card_linked_to_payments_profile", False)),
            "registration_fee_paid": bool(payments_info.get("registration_fee_paid", False)),
            "payment_method_status": "ready" if payments_info.get("card_linked_to_payments_profile") else "manual_required",
            "billing_profile_note": "Google Payments profile legal name/address must exactly match D&B (DUNS) profile.",
        },
    }
    record["readiness"] = _readiness_summary(record)
    return record


def _normalize_name(name: str) -> str:
    name = name.upper()
    name = re.sub(r"\b(LTD|LIMITED|PLC|LLP|CIC|CIO|INC|CORP|LLC|COMMUNITY INTEREST COMPANY)\b\.?", "", name)
    # Normalise & <-> AND (& has no word boundaries so handle separately)
    name = re.sub(r"\s*&\s*", " AND ", name)
    name = re.sub(r"[^A-Z0-9]+", "", name)
    return name


def _readiness_summary(record: dict) -> dict:
    checks = {
        # Google Account
        "gmail_owner_available": bool(record["account_owner"]["gmail_login"]),
        "gmail_2sv_enabled": bool(record["account_owner"]["2sv_enabled"]),
        # Company data
        "company_data_available": bool(record["organization"]["legal_name"] and record["organization"]["legal_address"]),
        "director_available": bool(record["authorized_representative"]["name"]),
        # DUNS
        "duns_available": bool(record["duns"]["duns_number"]),
        "duns_name_matches_ch": bool(record["duns"]["name_match"]),
        "duns_address_matches_ch": bool(record["duns"]["address_match"]),
        # Domain
        "domain_available": bool(record["domain"]["domain"]),
        # Developer email
        "developer_email_defined": bool(record["developer_contact"]["public_developer_email"]),
        "developer_email_forwarding_configured": record["developer_contact"]["email_forwarding"]["status"] == "configured",
        # Phones — phone is always assigned (may be dummy)
        "organization_phone_available": bool(record["organization"]["organization_phone"]),
        "contact_phone_otp_capable": bool(record["developer_contact"].get("phone_otp_capable", False)),
        "developer_phone_available": bool(record["developer_contact"].get("developer_phone", "")),
        # Certificate & documents
        "certificate_available": bool(record["documents"]["certificate_exists"]),
        "certificate_name_matches_duns": bool(record["duns"]["certificate_name_matches_duns"]),
        # DNS TXT (no website needed)
        "google_txt_token_added": record["dns_verification"]["status"] not in ("token_pending", "not_configured"),
        # Representative ID & proof of address
        "representative_id_collected": record["documents"]["representative_id_status"] == "ready",
        "proof_of_address_collected": record["documents"]["proof_of_address_status"] == "ready",
        # Payments
        "payment_card_linked": record["payments"]["payment_method_status"] == "ready",
    }
    blocking_missing = [name for name, ok in checks.items() if not ok]
    score = sum(1 for ok in checks.values() if ok)
    return {
        "checks": checks,
        "score": f"{score}/{len(checks)}",
        "ready_for_human_signup": not blocking_missing,
        "blocking_missing": blocking_missing,
    }


def write_readiness_sheet(workbook, results: list[dict]) -> None:
    """Add a Play Console Readiness sheet to an openpyxl workbook."""
    from openpyxl.styles import Alignment, Font, PatternFill

    if "Play Console Readiness" in workbook.sheetnames:
        del workbook["Play Console Readiness"]
    ws = workbook.create_sheet("Play Console Readiness")
    headers = [
        "No.", "Company Number", "Organization Legal Name", "Developer Display Name",
        "Gmail Owner", "2SV Enabled", "Authorized Representative", "Rep Role",
        "DUNS Number", "D&B Name", "Name Match", "Address Match",
        "Organization Phone", "Phone OTP Capable", "Phone Is Dummy",
        "Domain", "Org Website", "Developer Email", "Dev Email Forwarding",
        "Google TXT Status", "TXT Propagated",
        "Certificate", "Cert Name Matches DUNS", "Rep ID Status", "Proof of Address",
        "Card Linked", "Reg Fee Paid",
        "Score", "Missing Before Signup",
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color="23422F", end_color="23422F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, result in enumerate(results, 1):
        rec = build_readiness_record(result)
        missing = ", ".join(rec["readiness"]["blocking_missing"])
        ws.append([
            i,
            rec["company_number"],
            rec["organization"]["legal_name"],
            rec["organization"]["developer_display_name"],
            rec["account_owner"]["gmail_login"],
            "Yes" if rec["account_owner"]["2sv_enabled"] else "No",
            rec["authorized_representative"].get("display_name") or rec["authorized_representative"]["name"],
            rec["authorized_representative"]["role"],
            rec["duns"]["duns_number"],
            rec["duns"]["dnb_legal_name"],
            "Match" if rec["duns"]["name_match"] else f"Review: {rec['duns']['name_match_detail']}",
            "Match" if rec["duns"]["address_match"] else f"Review: {rec['duns']['address_match_detail']}",
            rec["organization"]["organization_phone"],
            "Yes" if rec["developer_contact"].get("phone_otp_capable") else "No (dummy)",
            "Yes" if rec["developer_contact"].get("phone_is_dummy") else "No",
            rec["domain"]["domain"],
            rec["organization"]["organization_website"],
            rec["developer_contact"]["public_developer_email"],
            rec["developer_contact"]["email_forwarding"]["status"],
            rec["dns_verification"]["status"],
            "Yes" if rec["dns_verification"]["propagated"] else "No",
            rec["documents"]["organization_document_status"],
            "Yes" if rec["duns"]["certificate_name_matches_duns"] else "Review",
            rec["documents"]["representative_id_status"],
            rec["documents"]["proof_of_address_status"],
            "Yes" if rec["payments"]["card_linked_to_payments_profile"] else "No",
            "Yes" if rec["payments"]["registration_fee_paid"] else "No",
            rec["readiness"]["score"],
            missing,
        ])

    for col in ws.columns:
        col_letter = col[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)
    ws.freeze_panes = "A2"


def export_dossier(results: list[dict], output_dir: str = DOSSIER_DIR) -> dict[str, str]:
    """Write Markdown and JSON dossiers for all results."""
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    records = [build_readiness_record(r) for r in results]
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = os.path.join(output_dir, f"play_console_readiness_{stamp}.json")
    md_path = os.path.join(output_dir, f"play_console_dossier_{stamp}.md")

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, ensure_ascii=False)

    with open(md_path, "w", encoding="utf-8") as f:
        f.write("# Google Play Console Organization Signup Dossier\n\n")
        f.write("This dossier is for human form filling only. Do not automate Google Play Console or Search Console.\n\n")
        for rec in records:
            _write_record_markdown(f, rec)
    return {"json": json_path, "markdown": md_path, "count": str(len(records))}


def _write_record_markdown(f, rec: dict) -> None:
    f.write(f"## {rec['organization']['legal_name']} ({rec['company_number']})\n\n")
    f.write(f"- Overall ready for human signup: {rec['readiness']['ready_for_human_signup']}\n")
    if rec["readiness"]["blocking_missing"]:
        f.write(f"- Missing before signup: {', '.join(rec['readiness']['blocking_missing'])}\n")
    f.write("\n")

    sections = [
        ("Account Owner", rec["account_owner"]),
        ("Organization", rec["organization"]),
        ("Contact Information (Private)", rec["contact"]),
        ("Authorized Representative", rec["authorized_representative"]),
        ("DUNS & Name Matching", rec["duns"]),
        ("Domain", rec["domain"]),
        ("Developer Contact", rec["developer_contact"]),
        ("DNS Verification (TXT Record)", rec["dns_verification"]),
        ("Documents", rec["documents"]),
        ("Payments", rec["payments"]),
    ]
    for title, data in sections:
        f.write(f"### {title}\n\n")
        _write_dict(f, data)
        f.write("\n")


def _write_dict(f, data: dict, prefix: str = "") -> None:
    for key, value in data.items():
        label = key.replace("_", " ").title()
        if isinstance(value, dict):
            f.write(f"- {prefix}{label}:\n")
            _write_dict(f, value, prefix="  ")
        elif isinstance(value, list):
            f.write(f"- {prefix}{label}: {', '.join(str(v) for v in value)}\n")
        else:
            f.write(f"- {prefix}{label}: {value}\n")
