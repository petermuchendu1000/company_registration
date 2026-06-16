"""
Email Pool Manager
==================
Manages a pool of Gmail accounts from emails.xlsx.
Assigns one email per company, extracts plausible names from email addresses,
and tracks usage to prevent reuse.

Usage:
    from email_pool import EmailPool
    pool = EmailPool()
    email, first_name, last_name, account_name = pool.assign_next()
    pool.mark_used(email, company_number="12345678", company_name="ACME LTD")
"""

import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR = os.path.dirname(__file__)
EMAILS_FILE = os.path.join(BASE_DIR, "emails.xlsx")


def extract_name_from_email(email):
    """Extract a plausible first and last name from an email address.

    Examples:
        faithkabebee@gmail.com   -> ("Faith", "Kabebee")
        john.doe@gmail.com       -> ("John", "Doe")
        jsmith123@gmail.com      -> ("Jsmith", "User")
        a@gmail.com              -> ("A", "User")

    For concatenated lowercase names (no separator) a table of common
    East-African/English first names is used to find the split boundary
    before falling back to a rough mid-point split.
    """
    local = email.split("@")[0]

    # Remove trailing digits
    local_clean = re.sub(r'\d+$', '', local)
    if not local_clean:
        local_clean = local

    # Remove leading/trailing underscores/dots/hyphens
    local_clean = local_clean.strip("._-")

    # Split on common separators
    parts = re.split(r'[._\-]+', local_clean)
    parts = [p for p in parts if p]

    if len(parts) >= 2:
        return parts[0].capitalize(), parts[-1].capitalize()

    if len(parts) == 1:
        name = parts[0]

        # Try camelCase split (works for mixed-case: johnOdhiambo)
        camel_parts = re.findall(r'[A-Z][a-z]+|[a-z]+(?=[A-Z]|$)', name)
        # Clean camel_parts: filter short fragments
        camel_parts = [p for p in camel_parts if len(p) >= 2]
        if len(camel_parts) >= 2:
            return camel_parts[0].capitalize(), camel_parts[-1].capitalize()

        # Try known first-name prefixes (East African + English common names)
        # Ordered longest-first so greedier matches win
        KNOWN_FIRST_NAMES = sorted([
            "faith", "grace", "mercy", "hope", "joy", "patience", "ruth",
            "esther", "mary", "alice", "jane", "anne", "rose", "lillian",
            "florence", "margaret", "patricia", "agnes", "beatrice", "doreen",
            "evelyn", "harriet", "irene", "josephine", "judith", "karen",
            "laura", "lydia", "miriam", "naomi", "pauline", "phoebe",
            "rachel", "rebecca", "sarah", "susan", "winnie", "abigail",
            "cynthia", "deborah", "diana", "elizabeth", "eunice", "gladys",
            "hellen", "immaculate", "jacqueline", "leah", "lilian", "linet",
            "monicah", "penina", "peris", "priscilla", "scholastica", "selina",
            "sheila", "tabitha", "teresia", "veronica", "violet", "vivian",
            "wambui", "wanjiku", "wairimu", "wangari", "achieng",
            # Male
            "john", "peter", "james", "robert", "david", "paul", "michael",
            "george", "thomas", "joseph", "charles", "henry", "william",
            "richard", "daniel", "andrew", "mark", "philip", "stephen",
            "patrick", "simon", "samuel", "benjamin", "timothy", "joshua",
            "gabriel", "jonathan", "christopher", "alexander", "emmanuel",
            "matthew", "kenneth", "brian", "dennis", "francis", "fredrick",
            "geoffrey", "gilead", "hassan", "ibrahim", "isaac", "isaiah",
            "julius", "kevin", "kipchoge", "leonard", "levi", "lewis",
            "linus", "martin", "moses", "nicholas", "noah", "obadiah",
            "oliver", "oscar", "owen", "rafael", "raymond", "rodgers",
            "ronald", "ruben", "seth", "silas", "silvester", "solomon",
            "terrence", "tobias", "victor", "vincent", "walter", "xavier",
        ], key=len, reverse=True)

        name_lower = name.lower()
        for first in KNOWN_FIRST_NAMES:
            if name_lower.startswith(first) and len(name_lower) > len(first) + 2:
                last = name_lower[len(first):]
                return first.capitalize(), last.capitalize()

        # If it looks like "initial + surname" (1 consonant + 4+ chars), treat
        # the first character as an initial — e.g. "cmuriithi" → ("C", "Muriithi")
        CONSONANTS = set("bcdfghjklmnpqrstvwxyz")
        if len(name) >= 5 and name[0] in CONSONANTS and name[1] in CONSONANTS:
            return name[0].upper(), name[1:].capitalize()

        # Fallback: rough mid-point split (better than no split)
        if len(name) > 4:
            mid = len(name) // 2
            return name[:mid].capitalize(), name[mid:].capitalize()

        return name.capitalize(), "User"

    return "Account", "User"


class EmailPool:
    """Manages a pool of Gmail accounts with assignment tracking."""

    def __init__(self, emails_file=None):
        self.emails_file = emails_file or EMAILS_FILE
        self._emails = []        # ordered list of all email addresses
        self._assignments = {}   # email -> {company_number, company_name}
        self._display_names = {} # email -> manually-set display name (overrides extraction)
        self._load()

    def _load(self):
        """Load emails and any existing assignments from the Excel file."""
        if not os.path.exists(self.emails_file):
            raise FileNotFoundError(f"Emails file not found: {self.emails_file}")

        wb = openpyxl.load_workbook(self.emails_file, read_only=True)
        ws = wb.active

        headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]

        email_col = headers.index("email") if "email" in headers else 0
        has_assigned = "assigned to (company number)" in headers
        cn_col = headers.index("assigned to (company number)") if has_assigned else None
        has_name_col = "assigned to (company name)" in headers
        name_col = headers.index("assigned to (company name)") if has_name_col else None
        # Optional column: human-readable display name that overrides extraction
        dn_col = headers.index("display name") if "display name" in headers else None

        for row in ws.iter_rows(min_row=2, values_only=True):
            email = row[email_col] if email_col < len(row) else None
            if not email or not isinstance(email, str):
                continue
            email = email.strip()
            if not email:
                continue

            self._emails.append(email)

            # Load optional manually-set display name
            if dn_col is not None and dn_col < len(row) and row[dn_col]:
                self._display_names[email] = str(row[dn_col]).strip()
            if cn_col is not None and cn_col < len(row) and row[cn_col]:
                company_number = str(row[cn_col]).strip()
                company_name = str(row[name_col]).strip() if name_col is not None and name_col < len(row) and row[name_col] else ""
                self._assignments[email] = {
                    "company_number": company_number,
                    "company_name": company_name,
                }

        wb.close()

    @property
    def total(self):
        return len(self._emails)

    @property
    def used_count(self):
        return len(self._assignments)

    @property
    def available_count(self):
        return self.total - self.used_count

    def _name_for_email(self, email):
        """Return (first_name, last_name, account_name) for an email address.

        Uses the "Display Name" column from emails.xlsx when available;
        falls back to extracting the name from the email address local part.
        """
        display = self._display_names.get(email)
        if display and display.strip():
            parts = display.strip().split(None, 1)
            first = parts[0]
            last  = parts[1] if len(parts) > 1 else ""
            account_name = display.strip()
        else:
            first, last = extract_name_from_email(email)
            account_name = f"{first} {last}".strip()
        return first, last, account_name

    def get_assigned_email(self, company_number):
        """Check if a company already has an assigned email."""
        for email, info in self._assignments.items():
            if info["company_number"] == str(company_number):
                first, last, account_name = self._name_for_email(email)
                return email, first, last, account_name
        return None, None, None, None

    def assign_next(self, company_number=None, company_name=""):
        """Assign the next available email. Returns (email, first_name, last_name, account_name).

        If company_number is provided and already has an assignment, returns that.
        """
        # Check if company already has an email
        if company_number:
            existing = self.get_assigned_email(company_number)
            if existing[0]:
                return existing

        # Find next unassigned email
        for email in self._emails:
            if email not in self._assignments:
                first, last, account_name = self._name_for_email(email)

                # Record assignment
                self._assignments[email] = {
                    "company_number": str(company_number) if company_number else "",
                    "company_name": company_name,
                }
                self._save()

                return email, first, last, account_name

        raise RuntimeError(f"No available emails! All {self.total} emails are assigned.")

    def _save(self):
        """Save the current state back to the Excel file."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Emails"

        headers = ["Email", "Display Name", "Extracted Name", "Assigned To (Company Number)", "Assigned To (Company Name)"]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="003078", end_color="003078", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for email in self._emails:
            first, last, account_name = self._name_for_email(email)
            assignment = self._assignments.get(email, {})
            auto_extracted = f"{first} {last}"
            ws.append([
                email,
                self._display_names.get(email, ""),   # Display Name (manual override)
                auto_extracted,                        # Extracted Name (auto)
                assignment.get("company_number", ""),
                assignment.get("company_name", ""),
            ])

        # Column widths
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 25  # Display Name
        ws.column_dimensions["C"].width = 25  # Extracted Name
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 45
        ws.freeze_panes = "A2"

        wb.save(self.emails_file)

    def status(self):
        """Print pool status."""
        print(f"Email Pool: {self.total} total, {self.used_count} assigned, {self.available_count} available")
        return {
            "total": self.total,
            "used": self.used_count,
            "available": self.available_count,
        }
