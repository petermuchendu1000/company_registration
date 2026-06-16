"""
Pipeline: Search Companies House -> Get company details -> Get DUNS -> Download certificate -> Register domain -> Assign email -> Save to Excel

Certificates saved to: ./certificates/{Company Name}/

Usage:
    python run_pipeline.py                         # Test with first company only
    python run_pipeline.py --count 5               # Process first 5 companies
    python run_pipeline.py --count 100             # Process all 100
    python run_pipeline.py --no-duns               # Skip DUNS (just details + cert)
    python run_pipeline.py --no-cert               # Skip certificate download
    python run_pipeline.py --no-domain             # Skip domain registration
    python run_pipeline.py --no-email              # Skip email assignment
    python run_pipeline.py --headless false        # Show browser for DUNS lookup
"""

import argparse
import re
import json
import os
import sys
import time
try:
    import requests
except ModuleNotFoundError:
    requests = None
from datetime import datetime
try:
    from dotenv import load_dotenv
except ModuleNotFoundError:
    def load_dotenv():
        return False

load_dotenv()

API_KEY = os.getenv("COMPANIES_HOUSE_API_KEY")
BASE_URL = "https://api.company-information.service.gov.uk"

# Output paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CERTS_BASE_DIR = os.path.join(BASE_DIR, "certificates")
OUTPUT_DIR = os.path.join(BASE_DIR, "pipeline_output")
EXCEL_FILE = os.path.join(OUTPUT_DIR, "companies_pipeline.xlsx")
PLAY_DOSSIER_DIR = os.path.join(OUTPUT_DIR, "play_console_dossiers")


def ch_get(path, params=None):
    """Companies House API GET with retry on 429 and network errors."""
    if requests is None:
        raise RuntimeError("The 'requests' package is required for live Companies House API calls.")
    for attempt in range(5):
        try:
            resp = requests.get(
                f"{BASE_URL}{path}",
                params=params,
                auth=(API_KEY, ""),
                timeout=15,
            )
        except (requests.ConnectionError, requests.Timeout) as e:
            wait = min(2 ** attempt * 5, 60)
            print(f"    Network error, retrying in {wait}s... ({e.__class__.__name__})")
            time.sleep(wait)
            continue
        if resp.status_code == 429:
            wait = min(2 ** attempt * 2, 30)
            print(f"    Rate limited, waiting {wait}s...")
            time.sleep(wait)
            continue
        resp.raise_for_status()
        return resp.json()
    raise Exception(f"Failed after 5 retries: {path}")


def ch_get_raw(url, headers=None, stream=False):
    """Raw GET for document downloads with 429 backoff."""
    for attempt in range(5):
        try:
            resp = requests.get(url, auth=(API_KEY, ""), headers=headers, stream=stream, timeout=30)
        except (requests.ConnectionError, requests.Timeout) as e:
            wait = min(2 ** attempt * 5, 60)
            print(f"    Network error, retrying in {wait}s... ({e.__class__.__name__})")
            time.sleep(wait)
            continue
        if resp.status_code == 429:
            wait = min(2 ** attempt * 2, 30)
            print(f"    Rate limited, waiting {wait}s...")
            time.sleep(wait)
            continue
        return resp
    return resp


# ============================================================
# Step 1: Search for companies (reuses the officer-name strategy)
# ============================================================

# Officer names per nationality (same as app.py)
OFFICER_NAME_TERMS = {
    "kenyan": [
        "mwangi", "odhiambo", "moraa", "wekesa", "mutua",
        "lenku", "kinoti", "kiptoo", "wanjiku", "onyango",
        "barasa", "nduku", "cheruiyot", "naserian", "muriithi",
        "mogaka", "achieng", "kamau", "wambua", "kaari",
        "kipchoge", "onsongo", "saitoti", "musyoka", "kerubo",
        "wanjala", "nyakundi", "ochieng", "wafula", "mutegi",
        "naisula", "wangari", "otieno", "lekatoo", "kiprotich",
        "kawira", "wanyonyi", "njoroge", "chebet", "mwikali",
    ],
}

MAX_OFFICERS = 3


def find_companies_by_nationality(nationality, count=1):
    """Find companies with directors of a given nationality, up to `count` results."""
    nat_lower = nationality.lower()
    officer_terms = OFFICER_NAME_TERMS.get(nat_lower, [])
    if not officer_terms:
        print(f"No officer name terms for nationality: {nationality}")
        return []

    companies = []
    seen = set()

    for name in officer_terms:
        if len(companies) >= count:
            break

        print(f"  Searching officer name: {name}...", end=" ", flush=True)

        for start in range(0, 60, 20):
            if len(companies) >= count:
                break

            try:
                data = ch_get("/search/officers", {"q": name, "items_per_page": 20, "start_index": start})
            except Exception as e:
                print(f"error: {e}")
                break

            officer_items = data.get("items", [])
            if not officer_items:
                break

            for officer in officer_items:
                if len(companies) >= count:
                    break

                appointments_link = officer.get("links", {}).get("self", "")
                if not appointments_link:
                    continue

                try:
                    appts = ch_get(appointments_link)
                except Exception:
                    continue

                for appt in appts.get("items", []):
                    if len(companies) >= count:
                        break

                    officer_nationality = (appt.get("nationality") or "").strip().lower()
                    if nat_lower not in officer_nationality:
                        continue

                    if appt.get("resigned_on"):
                        continue

                    company = appt.get("appointed_to", {})
                    cn = company.get("company_number", "")
                    status = (company.get("company_status") or "").lower()
                    if not cn or status != "active" or cn in seen:
                        continue

                    # Get company profile
                    try:
                        profile = ch_get(f"/company/{cn}")
                    except Exception:
                        continue

                    # Count active officers
                    active_count = 0
                    try:
                        officers_data = ch_get(f"/company/{cn}/officers")
                        for o in officers_data.get("items", []):
                            if not o.get("resigned_on"):
                                active_count += 1
                    except Exception:
                        active_count = 1

                    if active_count > MAX_OFFICERS:
                        continue

                    seen.add(cn)
                    companies.append({
                        "company_number": cn,
                        "company_name": profile.get("company_name", ""),
                        "search_name": name.title(),
                        "matched_officer": officer.get("title", name),
                    })
                    print(f"found {cn} ({profile.get('company_name', '')})")
                    break  # one company per officer

            if len(companies) >= count:
                break

        if len(companies) < count and not any(c.get("search_name", "").lower() == name for c in companies):
            print("no match")

    return companies


# ============================================================
# Step 2: Get detailed company data
# ============================================================

def get_company_details(company_number):
    """Get full company details from Companies House."""
    profile = ch_get(f"/company/{company_number}")
    officers_data = ch_get(f"/company/{company_number}/officers")

    addr = profile.get("registered_office_address", {})
    address_parts = [addr.get(k, "") for k in ["address_line_1", "address_line_2", "locality", "region", "postal_code", "country"] if addr.get(k)]

    # Active directors
    directors = []
    for o in officers_data.get("items", []):
        if not o.get("resigned_on") and "director" in (o.get("officer_role", "") or "").lower():
            directors.append({
                "name": o.get("name", ""),
                "nationality": o.get("nationality", ""),
                "appointed_on": o.get("appointed_on", ""),
            })

    sic_codes = profile.get("sic_codes", [])

    return {
        "company_number": company_number,
        "company_name": profile.get("company_name", ""),
        "company_status": profile.get("company_status", ""),
        "company_type": profile.get("type", ""),
        "date_of_creation": profile.get("date_of_creation", ""),
        "sic_codes": ", ".join(sic_codes),
        "address": ", ".join(address_parts),
        "address_line_1": addr.get("address_line_1", ""),
        "locality": addr.get("locality", ""),
        "postal_code": addr.get("postal_code", ""),
        "directors": directors,
        "director_names": "; ".join(d["name"] for d in directors),
        "director_nationalities": "; ".join(d["nationality"] for d in directors),
    }


# ============================================================
# Step 3: Get DUNS number
# ============================================================

def _get_dnb_session():
    """Get a requests session with D&B cookies for API access."""
    if not hasattr(_get_dnb_session, '_session'):
        s = requests.Session()
        s.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.8',
            'Accept-Encoding': 'gzip, deflate, br',
        })
        # Visit main page to get Akamai cookies
        s.get('https://www.dnb.co.uk/smb/duns/lookup.html', timeout=30)
        s.headers.update({
            'Accept': '*/*',
            'Referer': 'https://www.dnb.co.uk/smb/duns/lookup.html',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin',
        })
        _get_dnb_session._session = s
    return _get_dnb_session._session


def get_duns_number(company_number, headless=True):
    """Look up DUNS number via D&B's direct API endpoint."""
    api_url = 'https://www.dnb.co.uk/smb/duns/lookup/_jcr_content.criteriasearchservlet.json'
    params = {
        'unencryptedDUNS': 'true',
        'isDelisted': 'false',
        'countryISOAlpha2Code': 'GB',
        'registrationNumbers': company_number,
    }
    for attempt in range(3):
        try:
            session = _get_dnb_session()
            resp = session.get(api_url, params=params, timeout=30)
            resp.raise_for_status()
            data = resp.json()
            candidates = data.get('searchCandidates', [])
            if candidates:
                org = candidates[0].get('organization', {})
                duns = org.get('duns', '')
                name = org.get('primaryName', '')
                address = _format_dnb_address(org)
                return {
                    'duns_number': duns,
                    'status': 'found' if duns else 'not_found',
                    'company_name': name,
                    'dnb_name': name,
                    'dnb_address': address,
                }
            return {'duns_number': '', 'status': 'not_found'}
        except Exception as e:
            # Reset session on failure and retry
            if hasattr(_get_dnb_session, '_session'):
                del _get_dnb_session._session
            if attempt == 2:
                return {'duns_number': '', 'status': 'error', 'error': str(e)}
            time.sleep(2)


def _format_dnb_address(org):
    """Best-effort formatting for the varying D&B address JSON shapes."""
    addr = (
        org.get("primaryAddress")
        or org.get("registeredAddress")
        or org.get("businessAddress")
        or {}
    )
    if not isinstance(addr, dict):
        return ""
    parts = []
    street = addr.get("streetAddress") or {}
    if isinstance(street, dict):
        parts.extend(street.get(k, "") for k in ["line1", "line2"] if street.get(k))
    elif street:
        parts.append(str(street))
    for key in ["addressLocality", "addressRegion", "postalCode"]:
        val = addr.get(key)
        if isinstance(val, dict):
            val = val.get("name") or val.get("abbreviatedName")
        if val:
            parts.append(str(val))
    country = addr.get("addressCountry")
    if isinstance(country, dict):
        country = country.get("name") or country.get("isoAlpha2Code")
    if country:
        parts.append(str(country))
    return ", ".join(parts)


# ============================================================
# Step 4b: Register domain
# ============================================================

def _safe_folder_name(company_name):
    """Convert company name to a safe folder name."""
    name = re.sub(r'[<>:"/\\|?*]', '', company_name)
    name = name.strip('. ')
    return name or 'unknown'


def register_company_domain(company_name, details):
    """Find cheapest domain and register it for a company."""
    from namecheap_automation import (
        find_cheapest_domain,
        get_existing_account_contact,
        register_domain,
    )

    # Find cheapest available domain
    search_result = find_cheapest_domain(company_name)
    if "error" in search_result:
        return {"status": "error", "error": search_result["error"]}

    best = search_result["best"]
    domain = best["domain"]

    # Domain registration contact should match the configured Namecheap account
    # when possible. Fall back to Companies House/director details only if the
    # account profile cannot be read.
    registrant_info = get_existing_account_contact()
    contacts_source = registrant_info.get("source", "existing_namecheap_account")
    if "error" in registrant_info:
        directors = details.get("directors", [])
        if directors:
            # CH format: "SURNAME, Firstname"
            parts = directors[0]["name"].split(",", 1)
            if len(parts) == 2:
                last_name = parts[0].strip().title()
                first_name = parts[1].strip().split()[0].title() if parts[1].strip() else "Director"
            else:
                name_parts = directors[0]["name"].split()
                first_name = name_parts[0].title() if name_parts else "Director"
                last_name = name_parts[-1].title() if len(name_parts) > 1 else "Director"
        else:
            first_name = "Director"
            last_name = "Director"

        registrant_info = {
            "FirstName": first_name,
            "LastName": last_name,
            "OrganizationName": company_name,
            "Address1": details.get("address_line_1", "Companies House Default Address") or "Companies House Default Address",
            "City": details.get("locality", "London") or "London",
            "StateProvince": details.get("locality", "London") or "London",
            "PostalCode": details.get("postal_code", "EC1A 1BB") or "EC1A 1BB",
            "Country": "GB",
            "Phone": "+44.2070001000",
            "EmailAddress": f"admin@{domain}",
        }
        contacts_source = f"fallback_companies_house:{registrant_info.get('EmailAddress', '')}"

    reg_result = register_domain(domain, years=1, registrant_info=registrant_info)
    if "error" in reg_result:
        return {"status": "error", "error": reg_result["error"], "domain": domain}

    return {
        "status": "registered",
        "domain": domain,
        "charged": reg_result.get("charged", ""),
        "domain_id": reg_result.get("domain_id", ""),
        "order_id": reg_result.get("order_id", ""),
        "contacts_source": contacts_source,
        "alternatives": [a["domain"] for a in search_result.get("alternatives", [])],
    }


def setup_play_console_domain_assets(result, google_txt_token="", google_txt_host="@"):
    """
    Prepare Namecheap-side Play Console assets for a company:
      - dev@domain.com forwarding to the Gmail account owner
      - pending/configured Google TXT DNS record metadata

    This never interacts with Google. If no Google TXT token is supplied, DNS is
    left unchanged and the dossier records token_pending.
    """
    from namecheap_automation import (
        ensure_dev_email_forwarding,
        prepare_google_txt_record,
        set_default_dns,
    )
    from play_console_readiness import developer_email_for_domain

    domain = (result.get("domain") or {}).get("domain", "")
    account_email = (result.get("email") or {}).get("email", "")
    developer_email = developer_email_for_domain(domain)

    play = {
        "developer_email": developer_email,
        "developer_email_forwarding": {
            "status": "missing_domain_or_forward_target",
            "developer_email": developer_email,
            "forward_to": account_email,
        },
        "google_txt": {
            "status": "token_pending",
            "hostname": google_txt_host or "@",
            "value": google_txt_token or "",
        },
    }

    if not domain or not account_email:
        return play

    dns_result = set_default_dns(domain)
    play["namecheap_default_dns"] = dns_result

    forwarding_result = ensure_dev_email_forwarding(domain, account_email)
    play["developer_email_forwarding"] = forwarding_result

    txt_result = prepare_google_txt_record(
        domain,
        txt_value=google_txt_token,
        hostname=google_txt_host or "@",
    )
    play["google_txt"] = txt_result
    return play


# ============================================================
# Step 4: Download certificate of incorporation
# ============================================================

def download_certificate(company_number, output_dir):
    """Download the certificate of incorporation PDF."""
    os.makedirs(output_dir, exist_ok=True)

    data = ch_get(f"/company/{company_number}/filing-history", {"items_per_page": 100})

    # Look for incorporation filing
    doc_link = None
    for item in data.get("items", []):
        cat = (item.get("category") or "").lower()
        desc = (item.get("description") or "").lower()
        ftype = (item.get("type") or "").upper()
        if cat == "incorporation" or ftype == "NEWINC" or "incorporat" in desc:
            doc_link = item.get("links", {}).get("document_metadata")
            if doc_link:
                break

    # Fallback: earliest filing with a document
    if not doc_link:
        filings_with_docs = [
            item for item in data.get("items", [])
            if item.get("links", {}).get("document_metadata")
        ]
        if filings_with_docs:
            doc_link = filings_with_docs[-1].get("links", {}).get("document_metadata")

    if not doc_link:
        return None, "No documents available"

    # Build full URL
    if doc_link.startswith("/"):
        content_url = f"{BASE_URL}{doc_link}/content"
    else:
        content_url = f"{doc_link}/content"

    resp = ch_get_raw(content_url, headers={"Accept": "application/pdf"}, stream=True)
    if resp.status_code != 200:
        return None, f"HTTP {resp.status_code}"

    filename = f"Certificate_{company_number}.pdf"
    filepath = os.path.join(output_dir, filename)
    with open(filepath, "wb") as f:
        for chunk in resp.iter_content(chunk_size=8192):
            f.write(chunk)

    return filepath, None


# ============================================================
# Step 5: Save to Excel
def _write_results_to_excel(all_results, output_path):
    """Write results directly to Excel (no merge with existing)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Companies Pipeline"

    headers = [
        "No.", "Company Number", "Company Name", "Status", "Type",
        "Date of Creation", "SIC Codes", "Address",
        "Directors", "Director Nationalities",
        "DUNS Number", "DUNS Status", "DUNS Email Used",
        "D&B Legal Name", "D&B Address",
        "Certificate Downloaded", "Certificate Path",
        "Domain", "Domain Status", "Domain Cost",
        "Assigned Email", "Account Name",
        "Developer Email", "Dev Email Forwarding", "Google TXT Status",
        "Organization Phone", "Play Signup Missing",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="003078", end_color="003078", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, r in enumerate(all_results, 1):
        details = r.get("details", {})
        duns = r.get("duns", {})
        cert = r.get("certificate", {})
        domain = r.get("domain", {})
        email_info = r.get("email", {})
        from play_console_readiness import build_readiness_record
        readiness = build_readiness_record(r)

        ws.append([
            i,
            r.get("company_number", ""),
            details.get("company_name", ""),
            details.get("company_status", ""),
            details.get("company_type", ""),
            details.get("date_of_creation", ""),
            details.get("sic_codes", ""),
            details.get("address", ""),
            details.get("director_names", ""),
            details.get("director_nationalities", ""),
            duns.get("duns_number", ""),
            duns.get("status", ""),
            duns.get("temp_email", ""),
            duns.get("dnb_name", ""),
            duns.get("dnb_address", ""),
            "Yes" if cert.get("path") else "No",
            cert.get("path", cert.get("error", "")),
            domain.get("domain", ""),
            domain.get("status", ""),
            domain.get("charged", ""),
            email_info.get("email", ""),
            email_info.get("account_name", ""),
            readiness["developer_contact"]["public_developer_email"],
            readiness["developer_contact"]["email_forwarding"]["status"],
            readiness["dns_verification"]["status"],
            readiness["organization"]["organization_phone"],
            ", ".join(readiness["readiness"]["blocking_missing"]),
        ])

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    ws.freeze_panes = "A2"
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    from play_console_readiness import export_dossier, write_readiness_sheet
    write_readiness_sheet(wb, all_results)
    wb.save(output_path)
    dossier = export_dossier(all_results, PLAY_DOSSIER_DIR)
    print(f"\nExcel saved: {output_path} ({len(all_results)} total rows)")
    print(f"Play Console dossier: {dossier['markdown']}")


# ============================================================

def save_to_excel(results, output_path):
    """Save pipeline results to an Excel file. Merges with existing data."""
    try:
        import openpyxl
    except ImportError:
        print("Installing openpyxl...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl

    from openpyxl.styles import Font, PatternFill, Alignment

    # Load existing workbook if it exists
    existing_results = []
    if os.path.exists(output_path):
        existing_results = _load_existing_results(output_path)
        print(f"  Loaded {len(existing_results)} existing rows from Excel")

    # Merge: existing + new (no duplicates by company number)
    existing_numbers = {r.get("company_number") for r in existing_results}
    for r in results:
        if r.get("company_number") not in existing_numbers:
            existing_results.append(r)

    all_results = existing_results

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Companies Pipeline"

    # Headers
    headers = [
        "No.", "Company Number", "Company Name", "Status", "Type",
        "Date of Creation", "SIC Codes", "Address",
        "Directors", "Director Nationalities",
        "DUNS Number", "DUNS Status", "DUNS Email Used",
        "D&B Legal Name", "D&B Address",
        "Certificate Downloaded", "Certificate Path",
        "Domain", "Domain Status", "Domain Cost",
        "Assigned Email", "Account Name",
        "Developer Email", "Dev Email Forwarding", "Google TXT Status",
        "Organization Phone", "Play Signup Missing",
    ]
    ws.append(headers)

    # Style headers
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="003078", end_color="003078", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    for i, r in enumerate(all_results, 1):
        details = r.get("details", {})
        duns = r.get("duns", {})
        cert = r.get("certificate", {})

        domain = r.get("domain", {})
        email_info = r.get("email", {})
        from play_console_readiness import build_readiness_record
        readiness = build_readiness_record(r)

        ws.append([
            i,
            r.get("company_number", ""),
            details.get("company_name", ""),
            details.get("company_status", ""),
            details.get("company_type", ""),
            details.get("date_of_creation", ""),
            details.get("sic_codes", ""),
            details.get("address", ""),
            details.get("director_names", ""),
            details.get("director_nationalities", ""),
            duns.get("duns_number", ""),
            duns.get("status", ""),
            duns.get("temp_email", ""),
            duns.get("dnb_name", ""),
            duns.get("dnb_address", ""),
            "Yes" if cert.get("path") else "No",
            cert.get("path", cert.get("error", "")),
            domain.get("domain", ""),
            domain.get("status", ""),
            domain.get("charged", ""),
            email_info.get("email", ""),
            email_info.get("account_name", ""),
            readiness["developer_contact"]["public_developer_email"],
            readiness["developer_contact"]["email_forwarding"]["status"],
            readiness["dns_verification"]["status"],
            readiness["organization"]["organization_phone"],
            ", ".join(readiness["readiness"]["blocking_missing"]),
        ])

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Freeze header row
    ws.freeze_panes = "A2"

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    from play_console_readiness import export_dossier, write_readiness_sheet
    write_readiness_sheet(wb, all_results)
    wb.save(output_path)
    dossier = export_dossier(all_results, PLAY_DOSSIER_DIR)
    print(f"\nExcel saved: {output_path} ({len(all_results)} total rows)")
    print(f"Play Console dossier: {dossier['markdown']}")
    return output_path


def _load_existing_results(excel_path):
    """Load existing pipeline results from an Excel file."""
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb["Companies Pipeline"] if "Companies Pipeline" in wb.sheetnames else wb.active
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    header_map = {name: idx for idx, name in enumerate(header)}
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    wb.close()

    legacy = {
        "Company Number": 1,
        "Company Name": 2,
        "Status": 3,
        "Type": 4,
        "Date of Creation": 5,
        "SIC Codes": 6,
        "Address": 7,
        "Directors": 8,
        "Director Nationalities": 9,
        "DUNS Number": 10,
        "DUNS Status": 11,
        "DUNS Email Used": 12,
        "Certificate Downloaded": 13,
        "Certificate Path": 14,
        "Domain": 15,
        "Domain Status": 16,
        "Domain Cost": 17,
        "Assigned Email": 18,
        "Account Name": 19,
        "Developer Email": 20,
        "Dev Email Forwarding": 21,
        "Google TXT Status": 22,
        "Organization Phone": 23,
    }

    def cell(row, name, default=""):
        idx = header_map.get(name, legacy.get(name))
        if idx is None or len(row) <= idx or row[idx] is None:
            return default
        return row[idx]

    results = []
    for row in rows:
        if not row or not cell(row, "Company Number"):  # no company number
            continue
        results.append({
            "company_number": str(cell(row, "Company Number")),
            "details": {
                "company_name": cell(row, "Company Name"),
                "company_status": cell(row, "Status"),
                "company_type": cell(row, "Type"),
                "date_of_creation": cell(row, "Date of Creation"),
                "sic_codes": cell(row, "SIC Codes"),
                "address": cell(row, "Address"),
                "director_names": cell(row, "Directors"),
                "director_nationalities": cell(row, "Director Nationalities"),
            },
            "duns": {
                "duns_number": cell(row, "DUNS Number"),
                "status": cell(row, "DUNS Status"),
                "temp_email": cell(row, "DUNS Email Used"),
                "dnb_name": cell(row, "D&B Legal Name"),
                "dnb_address": cell(row, "D&B Address"),
            },
            "certificate": {
                "path": cell(row, "Certificate Path") if cell(row, "Certificate Downloaded") == "Yes" else "",
                "status": "downloaded" if cell(row, "Certificate Downloaded") == "Yes" else "",
                "error": cell(row, "Certificate Path") if cell(row, "Certificate Downloaded") != "Yes" else "",
            },
            "domain": {
                "domain": cell(row, "Domain"),
                "status": cell(row, "Domain Status"),
                "charged": cell(row, "Domain Cost"),
            },
            "email": {
                "email": cell(row, "Assigned Email"),
                "account_name": cell(row, "Account Name"),
            },
            "play_console": {
                "developer_email": cell(row, "Developer Email"),
                "developer_email_forwarding": {"status": cell(row, "Dev Email Forwarding")},
                "google_txt": {"status": cell(row, "Google TXT Status")},
            },
            "phone": {"phone_number": cell(row, "Organization Phone")},
        })
    return results


def load_existing_company_numbers():
    """Get set of company numbers already in the Excel file."""
    if not os.path.exists(EXCEL_FILE):
        return set()
    existing = _load_existing_results(EXCEL_FILE)
    return {r["company_number"] for r in existing}


# ============================================================
# Main Pipeline
# ============================================================

def run_pipeline(count=1, nationality="kenyan", skip_duns=False, skip_cert=False,
                 skip_domain=False, skip_email=False, headless=True,
                 google_txt_token="", google_txt_host="@"):
    """Run the full pipeline for `count` companies."""
    print("=" * 60)
    print(f"PIPELINE START - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Target: {count} companies | Nationality: {nationality}")
    print(f"DUNS: {'skip' if skip_duns else 'enabled'} | Cert: {'skip' if skip_cert else 'enabled'} | Domain: {'skip' if skip_domain else 'enabled'} | Email: {'skip' if skip_email else 'enabled'}")
    print("Google: no browser/API automation; only local dossier + optional Namecheap TXT/forwarding")
    print("=" * 60)

    # Initialize email pool
    email_pool = None
    if not skip_email:
        try:
            from email_pool import EmailPool
            email_pool = EmailPool()
            email_pool.status()
        except Exception as e:
            print(f"  Warning: Could not load email pool: {e}")
            email_pool = None

    # Step 1: Load existing companies to avoid duplicates
    existing_numbers = load_existing_company_numbers()
    if existing_numbers:
        print(f"\n  {len(existing_numbers)} companies already in Excel - will skip them")

    # Step 2: Find companies
    print(f"\n[1] SEARCHING for {count} companies with {nationality} directors...")
    companies = find_companies_by_nationality(nationality, count=count + len(existing_numbers))
    if not companies:
        print("No companies found. Exiting.")
        return

    # Filter out already-processed companies
    new_companies = [c for c in companies if c["company_number"] not in existing_numbers]
    new_companies = new_companies[:count]  # limit to requested count

    if not new_companies:
        print("All found companies are already in the Excel. Nothing to do.")
        return

    skipped = len(companies) - len(new_companies)
    if skipped > 0:
        print(f"  Skipped {skipped} companies already in Excel")
    print(f"\n  Processing {len(new_companies)} new companies.")

    results = []

    for idx, company in enumerate(new_companies, 1):
        cn = company["company_number"]
        print(f"\n{'='*60}")
        print(f"[Company {idx}/{len(new_companies)}] {cn} - {company.get('company_name', 'N/A')}")
        print(f"{'='*60}")

        result = {"company_number": cn}

        # Step 2: Get details
        print(f"  [2/6] Getting company details...")
        try:
            details = get_company_details(cn)
            result["details"] = details
            print(f"    Name: {details['company_name']}")
            print(f"    Address: {details['address']}")
            print(f"    Directors: {details['director_names']}")
            print(f"    SIC: {details['sic_codes']}")
        except Exception as e:
            print(f"    ERROR: {e}")
            result["details"] = {"error": str(e)}

        # Step 3: Get DUNS
        if skip_duns:
            print(f"  [3/6] DUNS: skipped")
            result["duns"] = {"status": "skipped"}
        else:
            print(f"  [3/6] Looking up DUNS number (API)...")
            try:
                duns_result = get_duns_number(cn, headless=headless)
                result["duns"] = {
                    "duns_number": duns_result.get("duns_number"),
                    "status": duns_result.get("status", "unknown"),
                    "temp_email": duns_result.get("temp_email", ""),
                    "dnb_name": duns_result.get("dnb_name") or duns_result.get("company_name", ""),
                    "dnb_address": duns_result.get("dnb_address", ""),
                    "message": duns_result.get("message", ""),
                }
                if duns_result.get("duns_number"):
                    print(f"    DUNS: {duns_result['duns_number']}")
                else:
                    print(f"    Status: {duns_result.get('status', 'unknown')} - {duns_result.get('message', '')}")
            except Exception as e:
                print(f"    ERROR: {e}")
                result["duns"] = {"status": "error", "error": str(e)}

        # Step 4: Download certificate
        company_name = result.get("details", {}).get("company_name", cn)
        company_folder = os.path.join(CERTS_BASE_DIR, _safe_folder_name(company_name))

        if skip_cert:
            print(f"  [4/6] Certificate: skipped")
            result["certificate"] = {"status": "skipped"}
        else:
            print(f"  [4/6] Downloading certificate of incorporation...")
            try:
                cert_path, cert_error = download_certificate(cn, company_folder)
                if cert_path:
                    result["certificate"] = {"path": cert_path, "status": "downloaded"}
                    print(f"    Saved: {cert_path}")
                else:
                    result["certificate"] = {"error": cert_error, "status": "failed"}
                    print(f"    Failed: {cert_error}")
            except Exception as e:
                print(f"    ERROR: {e}")
                result["certificate"] = {"error": str(e), "status": "error"}

        # Step 5: Register domain
        if skip_domain:
            print(f"  [5/6] Domain: skipped")
            result["domain"] = {"status": "skipped"}
        else:
            print(f"  [5/6] Registering domain...")
            try:
                domain_result = register_company_domain(company_name, result.get("details", {}))
                result["domain"] = domain_result
                if domain_result.get("status") == "registered":
                    print(f"    Domain: {domain_result['domain']} (${domain_result.get('charged', '?')})")
                else:
                    print(f"    Failed: {domain_result.get('error', 'unknown')}")
            except Exception as e:
                print(f"    ERROR: {e}")
                result["domain"] = {"status": "error", "error": str(e)}

        # Step 6: Assign email from pool
        if skip_email or email_pool is None:
            print(f"  [6/6] Email: skipped")
            result["email"] = {"status": "skipped"}
        else:
            print(f"  [6/6] Assigning email from pool...")
            try:
                email, first_name, last_name = email_pool.assign_next(
                    company_number=cn,
                    company_name=company_name,
                )
                account_name = f"{first_name} {last_name}"
                result["email"] = {
                    "email": email,
                    "first_name": first_name,
                    "last_name": last_name,
                    "account_name": account_name,
                    "status": "assigned",
                }
                print(f"    Email: {email}")
                print(f"    Account name: {account_name}")
            except Exception as e:
                print(f"    ERROR: {e}")
                result["email"] = {"status": "error", "error": str(e)}

        print(f"  [7/7] Preparing Play Console readiness record...")
        try:
            result["play_console"] = setup_play_console_domain_assets(
                result,
                google_txt_token=google_txt_token,
                google_txt_host=google_txt_host,
            )
            dev_email = result["play_console"].get("developer_email", "")
            fwd = result["play_console"].get("developer_email_forwarding", {}).get("status", "")
            txt = result["play_console"].get("google_txt", {}).get("status", "")
            print(f"    Developer email: {dev_email or 'missing domain'}")
            print(f"    Dev forwarding: {fwd}")
            print(f"    Google TXT: {txt}")
        except Exception as e:
            print(f"    ERROR: {e}")
            result["play_console"] = {"status": "error", "error": str(e)}

        results.append(result)

    # Save to Excel
    print(f"\n[6] Saving to Excel...")
    save_to_excel(results, EXCEL_FILE)

    # Summary
    print(f"\n{'='*60}")
    print("PIPELINE COMPLETE")
    print(f"{'='*60}")
    print(f"Companies processed: {len(results)}")
    duns_found = sum(1 for r in results if r.get("duns", {}).get("duns_number"))
    duns_submitted = sum(1 for r in results if r.get("duns", {}).get("status") == "submitted")
    certs_ok = sum(1 for r in results if r.get("certificate", {}).get("status") == "downloaded")
    emails_assigned = sum(1 for r in results if r.get("email", {}).get("status") == "assigned")
    print(f"DUNS obtained: {duns_found}")
    print(f"DUNS submitted (waiting): {duns_submitted}")
    domains_ok = sum(1 for r in results if r.get("domain", {}).get("status") == "registered")
    print(f"Certificates downloaded: {certs_ok}")
    print(f"Domains registered: {domains_ok}")
    print(f"Emails assigned: {emails_assigned}")
    print(f"Excel: {EXCEL_FILE}")
    print(f"Certificates: {CERTS_BASE_DIR}")

    return results


def refill_pipeline(skip_duns=False, skip_cert=False, skip_domain=False,
                    skip_email=False, headless=True, google_txt_token="",
                    google_txt_host="@"):
    """Re-process existing Excel entries that are missing details/DUNS/etc."""
    print("=" * 60)
    print(f"REFILL MODE - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    if not os.path.exists(EXCEL_FILE):
        print("No Excel file found. Run pipeline first.")
        return

    existing = _load_existing_results(EXCEL_FILE)
    print(f"  {len(existing)} companies in Excel")

    # Initialize email pool
    email_pool = None
    if not skip_email:
        try:
            from email_pool import EmailPool
            email_pool = EmailPool()
            email_pool.status()
        except Exception as e:
            print(f"  Warning: Could not load email pool: {e}")
            email_pool = None

    # Find entries needing work
    needs_details = [r for r in existing if not r.get("details", {}).get("company_name")]
    needs_duns = [r for r in existing if not skip_duns and r.get("details", {}).get("company_name") and not r.get("duns", {}).get("duns_number") and r.get("duns", {}).get("status") not in ("submitted",)]
    needs_cert = [r for r in existing if not skip_cert and r.get("details", {}).get("company_name") and not r.get("certificate", {}).get("path")]
    needs_email = [r for r in existing if not skip_email and r.get("details", {}).get("company_name") and not r.get("email", {}).get("email") and r.get("email", {}).get("status") != "skipped"]
    needs_play = [
        r for r in existing
        if r.get("domain", {}).get("domain")
        and r.get("email", {}).get("email")
        and r.get("play_console", {}).get("developer_email_forwarding", {}).get("status") != "configured"
    ]

    print(f"  Need details: {len(needs_details)}")
    print(f"  Need DUNS: {len(needs_duns)}")
    print(f"  Need certificates: {len(needs_cert)}")
    print(f"  Need email: {len(needs_email)}")
    print(f"  Need Play Console Namecheap assets: {len(needs_play)}")

    updated = 0

    # Refill details
    for i, r in enumerate(needs_details, 1):
        cn = r["company_number"]
        print(f"\n  [{i}/{len(needs_details)}] Getting details for {cn}...")
        try:
            details = get_company_details(cn)
            r["details"] = details
            print(f"    {details['company_name']} - {details['address']}")
            updated += 1
        except Exception as e:
            print(f"    ERROR: {e}")

    # Refill DUNS
    for i, r in enumerate(needs_duns, 1):
        cn = r["company_number"]
        name = r["details"].get("company_name", cn)
        print(f"\n  [{i}/{len(needs_duns)}] Getting DUNS for {cn} ({name})...")
        try:
            duns_result = get_duns_number(cn, headless=headless)
            r["duns"] = {
                "duns_number": duns_result.get("duns_number"),
                "status": duns_result.get("status", "unknown"),
                "temp_email": duns_result.get("temp_email", ""),
                "dnb_name": duns_result.get("dnb_name") or duns_result.get("company_name", ""),
                "dnb_address": duns_result.get("dnb_address", ""),
            }
            if duns_result.get("duns_number"):
                print(f"    DUNS: {duns_result['duns_number']}")
            else:
                print(f"    Status: {duns_result.get('status')}")
            updated += 1
        except Exception as e:
            print(f"    ERROR: {e}")

    # Refill certificates
    for i, r in enumerate(needs_cert, 1):
        cn = r["company_number"]
        name = r["details"].get("company_name", cn)
        company_folder = os.path.join(CERTS_BASE_DIR, _safe_folder_name(name))
        print(f"\n  [{i}/{len(needs_cert)}] Downloading certificate for {cn} ({name})...")
        try:
            filepath, error = download_certificate(cn, company_folder)
            if filepath:
                r["certificate"] = {"path": filepath, "status": "downloaded"}
                print(f"    Saved: {filepath}")
            else:
                r["certificate"] = {"status": "error", "error": error or "unknown"}
                print(f"    Error: {error}")
            updated += 1
        except Exception as e:
            print(f"    ERROR: {e}")

    # Refill emails
    for i, r in enumerate(needs_email, 1):
        if email_pool is None:
            break
        cn = r["company_number"]
        name = r["details"].get("company_name", cn)
        print(f"\n  [{i}/{len(needs_email)}] Assigning email for {cn} ({name})...")
        try:
            email, first_name, last_name = email_pool.assign_next(
                company_number=cn, company_name=name,
            )
            r["email"] = {
                "email": email,
                "first_name": first_name,
                "last_name": last_name,
                "account_name": f"{first_name} {last_name}",
                "status": "assigned",
            }
            print(f"    {email} ({first_name} {last_name})")
            updated += 1
        except Exception as e:
            print(f"    ERROR: {e}")

    # Refill Play Console Namecheap assets
    for i, r in enumerate(needs_play, 1):
        cn = r["company_number"]
        domain = r.get("domain", {}).get("domain", "")
        print(f"\n  [{i}/{len(needs_play)}] Preparing Play Console assets for {cn} ({domain})...")
        try:
            r["play_console"] = setup_play_console_domain_assets(
                r,
                google_txt_token=google_txt_token,
                google_txt_host=google_txt_host,
            )
            print(f"    Developer email: {r['play_console'].get('developer_email', '')}")
            print(f"    Forwarding: {r['play_console'].get('developer_email_forwarding', {}).get('status')}")
            print(f"    Google TXT: {r['play_console'].get('google_txt', {}).get('status')}")
            updated += 1
        except Exception as e:
            print(f"    ERROR: {e}")

    if updated > 0:
        print(f"\n  Saving {updated} updates...")
        # Write directly - don't use save_to_excel which re-loads and merges
        _write_results_to_excel(existing, EXCEL_FILE)
    else:
        print("\n  Nothing to update.")

    return existing


def generate_play_console_dossier_only():
    """Generate Play Console readiness workbook sheet + dossier from existing Excel only."""
    if not os.path.exists(EXCEL_FILE):
        print("No Excel file found. Run pipeline first.")
        return None

    existing = _load_existing_results(EXCEL_FILE)
    print(f"Loaded {len(existing)} existing companies from {EXCEL_FILE}")
    _write_results_to_excel(existing, EXCEL_FILE)
    return existing


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Companies Pipeline")
    parser.add_argument("--count", type=int, default=1, help="Number of companies to process")
    parser.add_argument("--nationality", default="kenyan", help="Officer nationality filter")
    parser.add_argument("--no-duns", action="store_true", help="Skip DUNS lookup")
    parser.add_argument("--no-cert", action="store_true", help="Skip certificate download")
    parser.add_argument("--no-domain", action="store_true", help="Skip domain registration")
    parser.add_argument("--no-email", action="store_true", help="Skip email assignment")
    parser.add_argument("--headless", default="false", help="Browser headless mode (true/false). D&B blocks headless.")
    parser.add_argument("--refill", action="store_true", help="Re-process existing entries missing data")
    parser.add_argument("--play-dossier-only", action="store_true", help="Only generate Play Console readiness sheet and dossier from existing Excel; no live APIs")
    parser.add_argument("--google-txt-token", default="", help="Optional Google TXT token copied manually from Google; applied only to Namecheap DNS")
    parser.add_argument("--google-txt-host", default="@", help="Hostname for Google TXT token, usually @")

    args = parser.parse_args()
    headless = args.headless.lower() != "false"

    if args.play_dossier_only:
        generate_play_console_dossier_only()
    elif args.refill:
        refill_pipeline(
            skip_duns=args.no_duns,
            skip_cert=args.no_cert,
            skip_domain=args.no_domain,
            skip_email=args.no_email,
            headless=headless,
            google_txt_token=args.google_txt_token,
            google_txt_host=args.google_txt_host,
        )
    else:
        run_pipeline(
            count=args.count,
            nationality=args.nationality,
            skip_duns=args.no_duns,
            skip_cert=args.no_cert,
            skip_domain=args.no_domain,
            skip_email=args.no_email,
            headless=headless,
            google_txt_token=args.google_txt_token,
            google_txt_host=args.google_txt_host,
        )
