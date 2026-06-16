"""
Email Pool Manager
==================
Manages a pool of Gmail accounts from emails.xlsx.
Assigns one email per company, extracts plausible names from email addresses,
and tracks usage to prevent reuse.

Usage:
    from email_pool import EmailPool
    pool = EmailPool()
    email, first_name, last_name = pool.assign_next()
    pool.mark_used(email, company_number="12345678", company_name="ACME LTD")
"""

import os
import re
import threading
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR = os.path.dirname(__file__)
EMAILS_FILE = os.path.join(BASE_DIR, "emails.xlsx")


def extract_name_from_email(email):
    """Extract a plausible first and last name from an email address.

    Examples:
        faithkabebee@gmail.com   -> ("Faith", "Kabebee")
        john.doe@gmail.com       -> ("John", "Doe")
        jsmith123@gmail.com      -> ("Jsmith", "User")
        a@gmail.com              -> ("A", "User")
    """
    local = email.split("@")[0]

    # Remove trailing digits
    local_clean = re.sub(r'\d+$', '', local)
    if not local_clean:
        local_clean = local

    # Remove leading/trailing underscores/dots/hyphens
    local_clean = local_clean.strip("._-")

    # Split on common separators
    parts = re.split(r'[._\-]+', local_clean)
    parts = [p for p in parts if p]

    if len(parts) >= 2:
        first_name = parts[0].capitalize()
        last_name = parts[-1].capitalize()
    elif len(parts) == 1:
        name = parts[0]
        # Try to split camelCase
        camel_parts = re.findall(r'[A-Z]?[a-z]+', name)
        if len(camel_parts) >= 2:
            first_name = camel_parts[0].capitalize()
            last_name = camel_parts[-1].capitalize()
        elif len(name) > 4:
            # Split roughly in half
            mid = len(name) // 2
            first_name = name[:mid].capitalize()
            last_name = name[mid:].capitalize()
        else:
            first_name = name.capitalize()
            last_name = "User"
    else:
        first_name = "Account"
        last_name = "User"

    return first_name, last_name


class EmailPool:
    """Manages a pool of Gmail accounts with assignment tracking."""

    def __init__(self, emails_file=None):
        self.emails_file = emails_file or EMAILS_FILE
        self._emails = []
        self._assignments = {}  # email -> {company_number, company_name}
        self._lock = threading.Lock()
        self._load()

    def _load(self):
        """Load emails and any existing assignments from the Excel file."""
        if not os.path.exists(self.emails_file):
            raise FileNotFoundError(f"Emails file not found: {self.emails_file}")

        wb = openpyxl.load_workbook(self.emails_file, read_only=True)
        ws = wb.active

        headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]

        email_col = headers.index("email") if "email" in headers else 0
        has_assigned = "assigned to (company number)" in headers
        cn_col = headers.index("assigned to (company number)") if has_assigned else None
        has_name_col = "assigned to (company name)" in headers
        name_col = headers.index("assigned to (company name)") if has_name_col else None

        for row in ws.iter_rows(min_row=2, values_only=True):
            email = row[email_col] if email_col < len(row) else None
            if not email or not isinstance(email, str):
                continue
            email = email.strip()
            if not email:
                continue

            self._emails.append(email)

            # Load existing assignment
            if cn_col is not None and cn_col < len(row) and row[cn_col]:
                company_number = str(row[cn_col]).strip()
                company_name = str(row[name_col]).strip() if name_col is not None and name_col < len(row) and row[name_col] else ""
                self._assignments[email] = {
                    "company_number": company_number,
                    "company_name": company_name,
                }

        wb.close()

    @property
    def total(self):
        return len(self._emails)

    @property
    def used_count(self):
        return len(self._assignments)

    @property
    def available_count(self):
        return self.total - self.used_count

    def get_assigned_email(self, company_number):
        """Check if a company already has an assigned email."""
        for email, info in self._assignments.items():
            if info["company_number"] == str(company_number):
                first, last = extract_name_from_email(email)
                return email, first, last
        return None, None, None

    def assign_next(self, company_number=None, company_name=""):
        """Assign the next available email. Returns (email, first_name, last_name).

        If company_number is provided and already has an assignment, returns that.
        """
        with self._lock:
            # Check if company already has an email
            if company_number:
                existing = self.get_assigned_email(company_number)
                if existing[0]:
                    return existing

            # Find next unassigned email
            for email in self._emails:
                if email not in self._assignments:
                    first_name, last_name = extract_name_from_email(email)

                    self._assignments[email] = {
                        "company_number": str(company_number) if company_number else "",
                        "company_name": company_name,
                    }
                    self._save()

                    return email, first_name, last_name

        raise RuntimeError(f"No available emails! All {self.total} emails are assigned.")

    def _save(self):
        """Save the current state back to the Excel file."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Emails"

        headers = ["Email", "Extracted Name", "Assigned To (Company Number)", "Assigned To (Company Name)"]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="003078", end_color="003078", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for email in self._emails:
            first, last = extract_name_from_email(email)
            assignment = self._assignments.get(email, {})
            ws.append([
                email,
                f"{first} {last}",
                assignment.get("company_number", ""),
                assignment.get("company_name", ""),
            ])

        # Column widths
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 45
        ws.freeze_panes = "A2"

        wb.save(self.emails_file)

    def status(self):
        """Print pool status."""
        print(f"Email Pool: {self.total} total, {self.used_count} assigned, {self.available_count} available")
        return {
            "total": self.total,
            "used": self.used_count,
            "available": self.available_count,
        }
