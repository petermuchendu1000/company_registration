"""
Wizard API routes — bolt onto the existing app.py Flask server.

Add to the bottom of app.py (before `if __name__ == "__main__":`)
or import via:  from wizard_routes import register_wizard_routes
                register_wizard_routes(app)
"""
from flask import Blueprint, request, jsonify
import os
import json
import sys
import threading
import subprocess
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

wizard = Blueprint("wizard", __name__)

EXCEL_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "pipeline_output", "companies_pipeline.xlsx",
)
MANUAL_INPUTS_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "pipeline_output", "manual_inputs.json",
)
MANUAL_INPUTS_FILE_GGPLAY = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "ggplay", "pipeline_output", "manual_inputs.json",
)

_PIPELINE_LOCK = threading.Lock()
_PIPELINE_RUN_STATE = {
    "process": None,
    "pid": None,
    "started_at": None,
    "finished_at": None,
    "exit_code": None,
    "command": None,
    "error": None,
    "log_path": None,
    "log_handle": None,
}


def _resolve_pipeline_script():
    root = os.path.dirname(os.path.abspath(__file__))
    ggplay_script = os.path.join(root, "ggplay", "run_pipeline.py")
    if os.path.exists(ggplay_script):
        return ggplay_script
    return os.path.join(root, "run_pipeline.py")


def _snapshot_pipeline_state_locked():
    proc = _PIPELINE_RUN_STATE.get("process")
    running = bool(proc and proc.poll() is None)

    if proc and not running and _PIPELINE_RUN_STATE.get("exit_code") is None:
        _PIPELINE_RUN_STATE["exit_code"] = proc.poll()
        _PIPELINE_RUN_STATE["finished_at"] = datetime.now().isoformat()
        _PIPELINE_RUN_STATE["process"] = None
        log_handle = _PIPELINE_RUN_STATE.get("log_handle")
        if log_handle:
            try:
                log_handle.flush()
                log_handle.close()
            except Exception:
                pass
            _PIPELINE_RUN_STATE["log_handle"] = None

    return {
        "running": running,
        "pid": _PIPELINE_RUN_STATE.get("pid"),
        "started_at": _PIPELINE_RUN_STATE.get("started_at"),
        "finished_at": _PIPELINE_RUN_STATE.get("finished_at"),
        "exit_code": _PIPELINE_RUN_STATE.get("exit_code"),
        "command": _PIPELINE_RUN_STATE.get("command"),
        "error": _PIPELINE_RUN_STATE.get("error"),
        "log_path": _PIPELINE_RUN_STATE.get("log_path"),
    }


def _latest_pipeline_log_path():
    logs_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pipeline_output", "logs")
    pattern = os.path.join(logs_dir, "pipeline_run_*.log")
    candidates = [p for p in glob.glob(pattern) if os.path.isfile(p)]
    if not candidates:
        return None
    return max(candidates, key=os.path.getmtime)

# ──────────────────────────────────────────────────────────────
# The canonical header row — must match run_pipeline.py + new cols
# ──────────────────────────────────────────────────────────────
HEADERS = [
    "No.", "Company Number", "Company Name", "Short Name",
    "Status", "Type", "Date of Creation", "SIC Codes", "Address",
    "Directors", "Director Nationalities",
    "DUNS Number", "DUNS Status", "DUNS Email Used",
    "Certificate Downloaded", "Certificate Path",
    "Domain", "Domain Status", "Domain Cost",
    "Assigned Email", "Account Name",
    "Company Status", "Company Notes", "Archived",
]


def _col_index(header_row, name):
    """Find column index by header name (case-insensitive)."""
    for i, h in enumerate(header_row):
        if h and str(h).strip().lower() == name.lower():
            return i
    return None


def _ensure_column(ws, header_name):
    """Ensure a column exists and return zero-based index."""
    header = [cell.value for cell in ws[1]]
    idx = _col_index(header, header_name)
    if idx is not None:
        return idx
    ws.cell(row=1, column=len(header) + 1, value=header_name)
    return len(header)


def _ensure_excel():
    """Create the pipeline Excel if it doesn't exist."""
    os.makedirs(os.path.dirname(EXCEL_FILE), exist_ok=True)
    if os.path.exists(EXCEL_FILE):
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Companies Pipeline"
    ws.append(HEADERS)
    # Style headers
    hdr_fill = PatternFill(start_color="003078", end_color="003078", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    wb.save(EXCEL_FILE)


def _load_manual_inputs_db():
    for path in [MANUAL_INPUTS_FILE, MANUAL_INPUTS_FILE_GGPLAY]:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                return data
    return {}


def _save_manual_inputs_db(db):
    for path in [MANUAL_INPUTS_FILE, MANUAL_INPUTS_FILE_GGPLAY]:
        try:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            with open(path, "w", encoding="utf-8") as f:
                json.dump(db, f, indent=2)
        except Exception:
            pass


# ──────────────────────────────────────────────────────────────
# POST /api/pipeline/add — add a company from the wizard
# ──────────────────────────────────────────────────────────────
@wizard.route("/api/pipeline/add", methods=["POST"])
def pipeline_add():
    data = request.get_json(force=True)

    cn = (data.get("company_number") or "").strip()
    if not cn:
        return jsonify({"success": False, "error": "company_number is required"}), 400

    _ensure_excel()

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Read header to find columns dynamically
    header = [cell.value for cell in ws[1]]

    # Check for duplicate
    col_cn = _col_index(header, "Company Number")
    if col_cn is not None:
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[col_cn].value and str(row[col_cn].value).strip() == cn:
                wb.close()
                return jsonify({"success": False, "error": f"Company {cn} already exists in the pipeline."}), 409

    # If "Short Name" column doesn't exist yet, add it
    col_sn = _col_index(header, "Short Name")
    if col_sn is None:
        # Insert after "Company Name"
        col_name = _col_index(header, "Company Name")
        insert_at = (col_name + 2) if col_name is not None else (len(header) + 1)
        ws.insert_cols(insert_at)
        ws.cell(row=1, column=insert_at, value="Short Name")
        # Re-read header
        header = [cell.value for cell in ws[1]]
        col_sn = _col_index(header, "Short Name")

    # Determine row number
    next_row = ws.max_row + 1
    row_num = next_row - 1  # 1-indexed company number

    def put(col_name, value):
        idx = _col_index(header, col_name)
        if idx is not None:
            ws.cell(row=next_row, column=idx + 1, value=value)

    put("No.", row_num)
    put("Company Number", cn)
    put("Company Name", data.get("company_name", ""))
    put("Short Name", data.get("short_name", ""))
    put("Status", data.get("status", ""))
    put("Type", data.get("type", ""))
    put("Date of Creation", data.get("date_of_creation", ""))
    put("SIC Codes", data.get("sic_codes", ""))
    put("Address", data.get("address", ""))
    put("Directors", data.get("directors", ""))
    put("Director Nationalities", data.get("nationalities", ""))
    put("DUNS Number", data.get("duns_number", ""))
    put("DUNS Status", "submitted" if data.get("duns_email") and not data.get("duns_number") else ("found" if data.get("duns_number") else ""))
    put("DUNS Email Used", data.get("duns_email", ""))
    put("Domain", data.get("domain", ""))
    put("Domain Status", "pending" if data.get("domain") else "")
    put("Assigned Email", data.get("email", ""))

    # Auto-width
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max_len + 2, 50)

    wb.save(EXCEL_FILE)
    wb.close()

    return jsonify({
        "success": True,
        "row": row_num,
        "company_number": cn,
        "message": f"Added {data.get('company_name', cn)} to pipeline."
    })


# ──────────────────────────────────────────────────────────────
# GET /api/pipeline/list — list all companies in the pipeline
# ──────────────────────────────────────────────────────────────
@wizard.route("/api/pipeline/list")
def pipeline_list():
    _ensure_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return jsonify([])

    header = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    companies = []
    for row in rows[1:]:
        entry = {}
        for i, val in enumerate(row):
            if i < len(header):
                entry[header[i]] = val
        companies.append(entry)

    return jsonify(companies)


# ──────────────────────────────────────────────────────────────
# GET /api/email-pool/status — email pool availability
# ──────────────────────────────────────────────────────────────
@wizard.route("/api/email-pool/status")
def email_pool_status():
    try:
        from email_pool import EmailPool
        pool = EmailPool()
        status = pool.status()
        available_emails = [e for e in pool._emails if e not in pool._assignments]
        return jsonify({
            "total": status["total"],
            "used": status["used"],
            "available": status["available"],
            "available_emails": available_emails[:50],  # cap at 50
        })
    except Exception as e:
        return jsonify({"error": str(e), "available_emails": []}), 500


# ──────────────────────────────────────────────────────────────
# GET /api/domains/check — check domain availability via Namecheap
# ──────────────────────────────────────────────────────────────
@wizard.route("/api/domains/check")
def domains_check():
    domains_str = request.args.get("domains", "").strip()
    if not domains_str:
        return jsonify({"error": "provide ?domains=example.co.uk"}), 400

    domain_list = [d.strip() for d in domains_str.split(",") if d.strip()]
    try:
        from namecheap_automation import check_domains
        results = check_domains(domain_list)
        return jsonify(results if isinstance(results, list) else [results])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ──────────────────────────────────────────────────────────────
# GET  /api/duns/lookup  — quick DUNS lookup
# POST /api/duns/request — submit DUNS request via stealth browser
# ──────────────────────────────────────────────────────────────
@wizard.route("/api/duns/lookup")
def duns_lookup():
    cn = request.args.get("company_number", "").strip()
    if not cn:
        return jsonify({"error": "company_number required"}), 400
    try:
        from duns_automation import lookup_duns
        result = lookup_duns(cn)
        return jsonify(result)
    except Exception as e:
        return jsonify({"found": False, "error": str(e)})


@wizard.route("/api/duns/request", methods=["POST"])
def duns_request():
    data = request.get_json(force=True)
    cn = data.get("company_number", "").strip()
    email = data.get("email", "").strip()
    if not cn or not email:
        return jsonify({"error": "company_number and email required"}), 400
    try:
        from duns_automation import stealth_request_duns
        result = stealth_request_duns(cn, headless=True)
        return jsonify(result)
    except Exception as e:
        return jsonify({"found": False, "error": str(e)})


# ──────────────────────────────────────────────────────────────
# Registration helper
# ──────────────────────────────────────────────────────────────


# ──────────────────────────────────────────────────────────────
# POST /api/domains/register — actually register via Namecheap
# ──────────────────────────────────────────────────────────────
@wizard.route("/api/domains/register", methods=["POST"])
def domains_register():
    data = request.get_json(force=True)
    domain = data.get("domain", "").strip()
    if not domain:
        return jsonify({"error": "domain is required"}), 400
    try:
        from namecheap_automation import register_domain
        result = register_domain(domain)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ──────────────────────────────────────────────────────────────
# POST /api/domains/find-cheapest — auto-find best domain for company
# ──────────────────────────────────────────────────────────────
@wizard.route("/api/domains/find-cheapest", methods=["POST"])
def domains_find_cheapest():
    data = request.get_json(force=True)
    company_name = data.get("company_name", "").strip()
    if not company_name:
        return jsonify({"error": "company_name required"}), 400
    try:
        from namecheap_automation import find_cheapest_domain
        result = find_cheapest_domain(company_name)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ──────────────────────────────────────────────────────────────
# POST /api/email-pool/assign — assign next available email
# ──────────────────────────────────────────────────────────────
@wizard.route("/api/email-pool/assign", methods=["POST"])
def email_pool_assign():
    data = request.get_json(force=True)
    company_number = data.get("company_number", "").strip()
    company_name = data.get("company_name", "")
    if not company_number:
        return jsonify({"error": "company_number required"}), 400
    try:
        from email_pool import EmailPool
        pool = EmailPool()
        # Check if already assigned
        existing = pool.get_assigned_email(company_number)
        if existing:
            return jsonify({"email": existing, "status": "already_assigned"})
        # Assign next
        result = pool.assign_next(company_number=company_number, company_name=company_name)
        if result:
            return jsonify({"email": result, "status": "assigned", "remaining": pool.available_count})
        return jsonify({"error": "No emails available in pool"}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ──────────────────────────────────────────────────────────────
# POST /api/certificate/download — download incorporation certificate
# ──────────────────────────────────────────────────────────────
@wizard.route("/api/certificate/download", methods=["POST"])
def certificate_download():
    data = request.get_json(force=True)
    company_number = data.get("company_number", "").strip()
    company_name = data.get("company_name", "")
    if not company_number:
        return jsonify({"error": "company_number required"}), 400
    try:
        import requests as http_requests
        from dotenv import load_dotenv
        load_dotenv()
        API_KEY = os.getenv("COMPANIES_HOUSE_API_KEY")
        BASE_URL = "https://api.company-information.service.gov.uk"
        # Get filing history
        resp = http_requests.get(
            f"{BASE_URL}/company/{company_number}/filing-history",
            params={"items_per_page": 100},
            auth=(API_KEY, ""),
        )
        resp.raise_for_status()
        filings = resp.json()
        # Find incorporation document
        doc_link = None
        for item in filings.get("items", []):
            cat = (item.get("category") or "").lower()
            desc = (item.get("description") or "").lower()
            ftype = (item.get("type") or "").upper()
            if cat == "incorporation" or ftype == "NEWINC" or "incorporat" in desc:
                links = item.get("links", {})
                doc_link = links.get("document_metadata")
                break
        if not doc_link:
            return jsonify({"status": "not_found", "message": "No incorporation filing found"})
        # Download PDF
        meta_resp = http_requests.get(f"https://api.company-information.service.gov.uk{doc_link}",
                                      auth=(API_KEY, ""))
        meta_resp.raise_for_status()
        meta = meta_resp.json()
        pdf_url = meta.get("links", {}).get("document")
        if not pdf_url:
            return jsonify({"status": "no_pdf", "message": "Document metadata found but no PDF link"})
        # Download the PDF
        pdf_resp = http_requests.get(pdf_url, auth=(API_KEY, ""),
                                     headers={"Accept": "application/pdf"})
        if pdf_resp.status_code == 200:
            safe_name = company_name.replace("/", "-").replace("\\", "-").strip() or company_number
            cert_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "certificates", safe_name)
            os.makedirs(cert_dir, exist_ok=True)
            cert_path = os.path.join(cert_dir, f"Certificate_{company_number}.pdf")
            with open(cert_path, "wb") as f:
                f.write(pdf_resp.content)
            return jsonify({"status": "downloaded", "path": cert_path, "size": len(pdf_resp.content)})
        return jsonify({"status": "download_failed", "http_status": pdf_resp.status_code})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ──────────────────────────────────────────────────────────────
# POST /api/domains/setup-email-forwarding
# ──────────────────────────────────────────────────────────────
@wizard.route("/api/domains/setup-email-forwarding", methods=["POST"])
def domains_email_forwarding():
    data = request.get_json(force=True)
    domain = data.get("domain", "").strip()
    forward_to = data.get("forward_to", "").strip()
    if not domain or not forward_to:
        return jsonify({"error": "domain and forward_to required"}), 400
    try:
        from namecheap_automation import set_email_forwarding
        result = set_email_forwarding(domain, "support", forward_to)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ──────────────────────────────────────────────────────────────
# POST /api/domains/apply-google-txt
# ──────────────────────────────────────────────────────────────
@wizard.route("/api/domains/apply-google-txt", methods=["POST"])
def domains_apply_google_txt():
    data = request.get_json(force=True)
    domain = (data.get("domain") or "").strip()
    txt_value = (data.get("txt_value") or "").strip()
    hostname = (data.get("hostname") or "@").strip() or "@"
    company_number = (data.get("company_number") or "").strip()

    if not domain:
        return jsonify({"error": "domain is required"}), 400
    if not txt_value:
        return jsonify({"error": "txt_value is required"}), 400

    try:
        from namecheap_automation import set_default_dns, add_txt_record

        dns_result = set_default_dns(domain)
        if "error" in dns_result:
            return jsonify({"error": dns_result["error"], "stage": "set_default_dns"}), 400

        txt_result = add_txt_record(domain, txt_value, hostname=hostname)
        if "error" in txt_result:
            return jsonify({"error": txt_result["error"], "stage": "add_txt_record"}), 400

        # Best-effort write-back to pipeline sheet if company exists.
        updated_excel = False
        excel_warning = ""
        if company_number and os.path.exists(EXCEL_FILE):
            try:
                wb = openpyxl.load_workbook(EXCEL_FILE)
                ws = wb.active
                header = [cell.value for cell in ws[1]]
                col_cn = _col_index(header, "Company Number")
                row_found = False
                if col_cn is not None:
                    col_txt_status = _ensure_column(ws, "Google TXT Status")
                    col_domain = _ensure_column(ws, "Domain")
                    for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        cell_cn = r[col_cn]
                        if cell_cn.value and str(cell_cn.value).strip() == company_number:
                            ws.cell(row=cell_cn.row, column=col_domain + 1, value=domain)
                            ws.cell(row=cell_cn.row, column=col_txt_status + 1, value="configured")
                            row_found = True
                            break
                    wb.save(EXCEL_FILE)
                    updated_excel = row_found
                wb.close()
            except Exception as e:
                excel_warning = f"TXT applied, but could not update Excel: {e}"

        return jsonify({
            "success": True,
            "domain": domain,
            "hostname": hostname,
            "txt_value": txt_value,
            "dns": dns_result,
            "txt": txt_result,
            "excel_updated": updated_excel,
            "warning": excel_warning,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ──────────────────────────────────────────────────────────────
# GET/POST /api/pipeline/manual-inputs/<company_number>
# ──────────────────────────────────────────────────────────────
@wizard.route("/api/pipeline/manual-inputs/<company_number>", methods=["GET"])
def pipeline_manual_inputs_get(company_number):
    try:
        db = _load_manual_inputs_db()
        return jsonify(db.get(company_number, {}))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@wizard.route("/api/pipeline/manual-inputs/<company_number>", methods=["POST"])
def pipeline_manual_inputs_save(company_number):
    try:
        payload = request.get_json(force=True) or {}

        manual = {
            "representative_id": {
                "full_name_as_on_id": (payload.get("representative_id", {}).get("full_name_as_on_id") or "").strip(),
                "personal_address": (payload.get("representative_id", {}).get("personal_address") or "").strip(),
                "id_type": (payload.get("representative_id", {}).get("id_type") or "").strip(),
                "id_country": (payload.get("representative_id", {}).get("id_country") or "").strip(),
                "id_expiry_date": (payload.get("representative_id", {}).get("id_expiry_date") or "").strip(),
                "id_image_front": (payload.get("representative_id", {}).get("id_image_front") or "").strip(),
                "id_image_back": (payload.get("representative_id", {}).get("id_image_back") or "").strip(),
                "id_quality_check": (payload.get("representative_id", {}).get("id_quality_check") or "pending").strip(),
            }
        }

        db = _load_manual_inputs_db()
        db[company_number] = manual
        _save_manual_inputs_db(db)

        excel_updated = False
        if os.path.exists(EXCEL_FILE):
            try:
                wb = openpyxl.load_workbook(EXCEL_FILE)
                ws = wb.active
                header = [cell.value for cell in ws[1]]
                col_cn = _col_index(header, "Company Number")
                if col_cn is not None:
                    col_rep = _ensure_column(ws, "Rep ID Status")

                    rep_ready = bool(manual["representative_id"].get("id_image_front")) and bool(manual["representative_id"].get("id_image_back"))

                    for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        cn_cell = r[col_cn]
                        if cn_cell.value and str(cn_cell.value).strip() == company_number:
                            ws.cell(row=cn_cell.row, column=col_rep + 1, value="ready" if rep_ready else "manual_required")
                            excel_updated = True
                            break
                    wb.save(EXCEL_FILE)
                wb.close()
            except Exception:
                excel_updated = False

        return jsonify({"success": True, "company_number": company_number, "excel_updated": excel_updated})
    except Exception as e:
        return jsonify({"error": str(e)}), 500





@wizard.route("/download")
def download_file():
    """Download a file from the pipeline output directory."""
    from flask import send_file, abort
    
    file_path = request.args.get("path", "").strip()
    if not file_path:
        return abort(400)

    # Security: only allow files under pipeline_output
    base_dir = os.path.join(os.path.dirname(__file__), "pipeline_output")
    abs_path = os.path.abspath(file_path)
    
    if not abs_path.startswith(os.path.abspath(base_dir)):
        return abort(403)

    if not os.path.exists(abs_path):
        return abort(404)

    return send_file(abs_path, as_attachment=True)


@wizard.route("/api/pipeline/retry/<company_number>", methods=["POST"])
def pipeline_retry(company_number):
    """
    Retry a failed company (re-trigger pipeline for a single company).
    Returns a job ID or status message.
    """
    if not company_number:
        return jsonify({"error": "company_number is required"}), 400

    try:
        from ggplay.run_pipeline import _load_existing_results, save_to_excel
        
        if not os.path.exists(EXCEL_FILE):
            return jsonify({"error": "Pipeline has not run yet"}), 400

        existing = _load_existing_results(EXCEL_FILE)
        company = next((c for c in existing if c.get("company_number") == company_number), None)
        
        if not company:
            return jsonify({"error": f"Company {company_number} not found"}), 404

        # Reset error states
        for key in ["certificate", "domain", "email", "apk_generation"]:
            if key in company:
                company[key] = {k: v for k, v in company[key].items() if k != "error"}

        # Save back to Excel
        save_to_excel([company], EXCEL_FILE)

        return jsonify({
            "success": True,
            "company_number": company_number,
            "message": "Company queued for retry. Pipeline will process on next run.",
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@wizard.route("/api/pipeline/run", methods=["POST"])
def pipeline_run_start():
    """Start a background pipeline run from dashboard action controls."""
    data = request.get_json(silent=True) or {}
    try:
        count = int(data.get("count", 1))
    except (TypeError, ValueError):
        return jsonify({"error": "count must be a valid integer"}), 400

    if count < 1 or count > 500:
        return jsonify({"error": "count must be between 1 and 500"}), 400

    nationality = (data.get("nationality") or "kenyan").strip().lower()
    if not nationality:
        nationality = "kenyan"

    with _PIPELINE_LOCK:
        state = _snapshot_pipeline_state_locked()
        if state["running"]:
            return jsonify({
                "error": "Pipeline is already running",
                "pid": state.get("pid"),
                "started_at": state.get("started_at"),
            }), 409

        script_path = _resolve_pipeline_script()
        if not os.path.exists(script_path):
            return jsonify({"error": f"Pipeline script not found: {script_path}"}), 500

        cmd = [
            sys.executable,
            "-u",
            script_path,
            "--count", str(count),
            "--nationality", nationality,
        ]

        logs_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pipeline_output", "logs")
        os.makedirs(logs_dir, exist_ok=True)
        log_file_name = f"pipeline_run_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        log_file_path = os.path.join(logs_dir, log_file_name)

        log_handle = None

        try:
            log_handle = open(log_file_path, "a", encoding="utf-8", buffering=1)
            log_handle.write(f"[launcher] starting at {datetime.now().isoformat()}\n")
            log_handle.write(f"[launcher] command: {' '.join(cmd)}\n")
            log_handle.write(f"[launcher] count={count} nationality={nationality}\n")
            env = os.environ.copy()
            env["PYTHONUNBUFFERED"] = "1"
            env["PYTHONIOENCODING"] = "utf-8"
            proc = subprocess.Popen(
                cmd,
                cwd=os.path.dirname(os.path.abspath(__file__)),
                stdout=log_handle,
                stderr=subprocess.STDOUT,
                env=env,
            )
        except Exception as e:
            if log_handle:
                try:
                    log_handle.close()
                except Exception:
                    pass
            return jsonify({"error": f"Failed to start pipeline: {e}"}), 500

        _PIPELINE_RUN_STATE["process"] = proc
        _PIPELINE_RUN_STATE["pid"] = proc.pid
        _PIPELINE_RUN_STATE["started_at"] = datetime.now().isoformat()
        _PIPELINE_RUN_STATE["finished_at"] = None
        _PIPELINE_RUN_STATE["exit_code"] = None
        _PIPELINE_RUN_STATE["command"] = " ".join(cmd)
        _PIPELINE_RUN_STATE["error"] = None
        _PIPELINE_RUN_STATE["log_path"] = log_file_path
        _PIPELINE_RUN_STATE["log_handle"] = log_handle

        return jsonify({
            "success": True,
            "message": "Pipeline started",
            "pid": proc.pid,
            "started_at": _PIPELINE_RUN_STATE["started_at"],
            "count": count,
            "nationality": nationality,
            "log_path": log_file_path,
        })


@wizard.route("/api/pipeline/run/status")
def pipeline_run_status():
    """Get latest background pipeline run status."""
    with _PIPELINE_LOCK:
        state = _snapshot_pipeline_state_locked()
    return jsonify(state)


@wizard.route("/api/pipeline/run/logs")
def pipeline_run_logs():
    """Return tail of pipeline run logs for live dashboard viewing."""
    try:
        tail_lines = int(request.args.get("tail", 300))
    except (TypeError, ValueError):
        tail_lines = 300

    if tail_lines < 50:
        tail_lines = 50
    if tail_lines > 2000:
        tail_lines = 2000

    with _PIPELINE_LOCK:
        state = _snapshot_pipeline_state_locked()
        log_path = _PIPELINE_RUN_STATE.get("log_path")
        log_handle = _PIPELINE_RUN_STATE.get("log_handle")
        if log_handle:
            try:
                log_handle.flush()
            except Exception:
                pass
        if not log_path:
            log_path = _latest_pipeline_log_path()
            if not log_path:
                return jsonify({"running": state.get("running", False), "log": "", "line_count": 0})

        logs_base = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "pipeline_output", "logs"))
        abs_log_path = os.path.abspath(log_path)
        if not abs_log_path.startswith(logs_base):
            return jsonify({"error": "invalid log path"}), 403

    if not os.path.exists(abs_log_path):
        return jsonify({"running": state.get("running", False), "log": "", "line_count": 0})

    try:
        with open(abs_log_path, "r", encoding="utf-8", errors="replace") as f:
            lines = f.readlines()
    except Exception as e:
        return jsonify({"error": f"failed reading logs: {e}"}), 500

    tail = lines[-tail_lines:]
    return jsonify({
        "running": state.get("running", False),
        "log": "".join(tail),
        "line_count": len(lines),
        "tail_lines": len(tail),
        "log_path": abs_log_path,
    })





@wizard.route("/api/company/<cn>/files")
def company_files_api(cn):
    """File tree for a single company (only key files)."""
    import re as _re
    ROOT = os.path.dirname(os.path.abspath(__file__))
    COMP_DIR = os.path.join(ROOT, "pipeline_output", "companies")

    name = ""
    if os.path.exists(EXCEL_FILE):
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
            ws = wb.active
            hdr = [str(c.value or "").strip().lower() for c in ws[1]]
            cn_c = next((i for i, h in enumerate(hdr) if h == "company number"), None)
            nm_c = next((i for i, h in enumerate(hdr) if h == "company name"), None)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or cn_c is None: continue
                if str(row[cn_c] or "").strip().zfill(8) == cn.zfill(8):
                    name = str(row[nm_c] or "").strip() if nm_c is not None else ""
                    break
            wb.close()
        except Exception:
            pass

    safe = _re.sub(r'[\\/:*?"<>|]', "", name).strip()
    folder = f"{cn.zfill(8)} - {safe}" if safe else cn.zfill(8)
    cdir = os.path.join(COMP_DIR, folder)
    sections = []

    # Certificate — only _cert.pdf (skip the large original)
    cert_d = os.path.join(cdir, "certificate")
    cert_files = []
    if os.path.isdir(cert_d):
        for f in sorted(os.listdir(cert_d)):
            fp = os.path.join(cert_d, f)
            if os.path.isfile(fp):
                cert_files.append({"name": f, "size": os.path.getsize(fp), "path": fp})
    sections.append({"name": "certificate", "files": cert_files})

    # Director ID — front, back, combined
    id_d = os.path.join(cdir, "director_id")
    id_files = []
    for f in ["ID Front.jpeg", "ID Back.jpeg", "ID_Combined.jpeg"]:
        fp = os.path.join(id_d, f)
        if os.path.exists(fp):
            id_files.append({"name": f, "size": os.path.getsize(fp), "path": fp})
    sections.append({"name": "director_id", "files": id_files})

    # App — APK and AAB only
    art_d = os.path.join(cdir, "app", "artifacts")
    app_files = []
    if os.path.isdir(art_d):
        for f in sorted(os.listdir(art_d)):
            if f.endswith((".apk", ".aab")):
                fp = os.path.join(art_d, f)
                app_files.append({"name": f, "size": os.path.getsize(fp), "path": fp})
    sections.append({"name": "app", "files": app_files})

    # Payment profile (check subdirectory first, then root)
    pp_files = []
    for pp_path in [
        os.path.join(cdir, "payment_profile", "payment_profile.txt"),
        os.path.join(cdir, "payment_profile.txt"),
    ]:
        if os.path.exists(pp_path):
            pp_files.append({"name": "payment_profile.txt", "size": os.path.getsize(pp_path), "path": pp_path})
            break
    sections.append({"name": "payment_profile", "files": pp_files})

    return jsonify({"cn": cn, "name": name, "sections": sections})


def register_wizard_routes(app):
    """Call this from app.py to mount all wizard endpoints."""
    app.register_blueprint(wizard)
